"""openpyxl helpers shared by the sanitizer, the validator and the exporter.

Three things live here because all three consumers need them and none of them should own
them:

* **Read-only loading.** ``openpyxl.load_workbook(read_only=True)`` streams rows instead
  of materialising the whole sheet, which is what keeps a ten-thousand-row workbook
  practical. Formulas cannot be read in that mode, so a second loader exists for the
  cases that genuinely need them, and the choice is explicit at every call site rather
  than a default nobody notices.
* **Atomic writing.** A workbook is written to a temporary file beside its destination
  and then replaced in one ``os.replace``. A run that fails half-way leaves the previous
  artifact untouched rather than a truncated file that still opens.
* **Styling.** The governed workbook look -- title, key/value blocks, styled tables,
  frozen headers, filters, currency and date formats, controlled widths -- expressed once
  so the exporter reproduces the sanitized artifact's contract rather than approximating
  it.

No macro is ever written or executed: ``.xlsx`` has no macro storage, and the exporter
never emits ``.xlsm``.
"""

from __future__ import annotations

import os
import tempfile
import warnings
from collections.abc import Iterator, Sequence
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Final

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from arpi.exceptions import ValidationError

__all__ = [
    "CURRENCY_FORMAT",
    "DATE_FORMAT",
    "INTEGER_FORMAT",
    "SheetData",
    "add_condition_chart",
    "add_top_model_chart",
    "atomic_workbook_save",
    "column_widths",
    "open_read_only",
    "open_with_formulas",
    "read_sheet_rows",
    "style_key_value_block",
    "style_table",
    "write_notice",
    "write_title",
]

#: Number formats used across every governed workbook. Declared once so the exporter and
#: the workbook tests compare against the same strings.
CURRENCY_FORMAT: Final = '"$"#,##0'
INTEGER_FORMAT: Final = "#,##0"
PERCENT_FORMAT: Final = "0.0%"
DATE_FORMAT: Final = "yyyy-mm-dd"

_TITLE_FONT: Final = Font(size=14, bold=True, color="1F2933")
_HEADING_FONT: Final = Font(size=11, bold=True, color="1F2933")
_LABEL_FONT: Final = Font(bold=True, color="3E4C59")
_NOTICE_FONT: Final = Font(size=9, italic=True, color="52606D")
_HEADER_FILL: Final = PatternFill("solid", fgColor="1F2933")
_HEADER_FONT: Final = Font(bold=True, color="FFFFFF")
_THIN: Final = Side(style="thin", color="CBD2D9")
_CELL_BORDER: Final = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


@contextmanager
def _quiet_openpyxl() -> Iterator[None]:
    """Silence openpyxl's "unsupported extension" notices for the duration of a load.

    A workbook produced by Excel carries extension parts -- conditional-formatting data
    bars, slicer caches -- that openpyxl cannot represent and drops on read. It says so
    with a ``UserWarning``, which is correct and also actionable by nobody: ARPI reads
    these files, it does not round-trip them, so nothing is lost by the drop.

    The project runs pytest with ``filterwarnings = ["error"]``, so leaving the notice
    unfiltered would turn "this real workbook contains a data bar" into a test failure.
    The filter is scoped to that one message text and to ``UserWarning``, so any other
    warning openpyxl raises still reaches the caller -- and it is scoped to the duration
    of a single load, so it cannot leak into the rest of the process.
    """
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r".*extension is not supported and will be removed.*",
            category=UserWarning,
        )
        yield


@dataclass(frozen=True, slots=True)
class SheetData:
    """A sheet reduced to its header row and its data rows.

    Attributes:
        title: Sheet name.
        headers: Row 1, as read.
        rows: Every subsequent row, trimmed to the header width.
    """

    title: str
    headers: tuple[str, ...]
    rows: tuple[tuple[Any, ...], ...]

    @property
    def row_count(self) -> int:
        """Number of data rows, excluding the header."""
        return len(self.rows)


@contextmanager
def open_read_only(path: Path) -> Iterator[Any]:
    """Open a workbook in streaming read-only mode and close it deterministically.

    Args:
        path: Workbook to open.

    Yields:
        The opened workbook.

    Raises:
        ValidationError: If the file is absent or is not a readable workbook. The message
            names the path, which is a repository path and never a source identifier.
    """
    if not path.is_file():
        raise ValidationError(f"Workbook not found: {path}.", field="input")
    # The filter wraps the whole block, not only the load. In read-only mode openpyxl
    # parses a sheet lazily while its rows are being iterated, so the extension notice is
    # raised inside the caller's `with`, not at load time.
    with _quiet_openpyxl():
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as error:  # noqa: BLE001 - openpyxl raises a wide range of types
            raise ValidationError(
                f"{path.name} could not be opened as an .xlsx workbook "
                f"({type(error).__name__}).",
                field="input",
            ) from error
        try:
            yield workbook
        finally:
            workbook.close()


@contextmanager
def open_with_formulas(path: Path) -> Iterator[Any]:
    """Open a workbook so that formula text is visible.

    Read-only mode cannot return formulas, so this loader materialises the sheets. It is
    used only where formulas are the thing being checked -- the validator's
    formula-presence rule and the workbook tests -- never on the row-reading path.

    Args:
        path: Workbook to open.

    Yields:
        The opened workbook.

    Raises:
        ValidationError: If the file is absent or unreadable.
    """
    if not path.is_file():
        raise ValidationError(f"Workbook not found: {path}.", field="input")
    with _quiet_openpyxl():
        try:
            workbook = load_workbook(path, read_only=False, data_only=False)
        except Exception as error:  # noqa: BLE001 - openpyxl raises a wide range of types
            raise ValidationError(
                f"{path.name} could not be opened as an .xlsx workbook "
                f"({type(error).__name__}).",
                field="input",
            ) from error
        try:
            yield workbook
        finally:
            workbook.close()


def read_sheet_rows(workbook: Any, title: str) -> SheetData:
    """Read one sheet's header row and data rows.

    Trailing rows that are entirely empty are dropped: Excel routinely reports a
    ``max_row`` beyond the last populated row, and counting those as data would make a
    row-count assertion depend on how the file was last saved.

    Args:
        workbook: An open workbook.
        title: Sheet name.

    Returns:
        The sheet's data.

    Raises:
        ValidationError: If the sheet is absent.
    """
    if title not in workbook.sheetnames:
        raise ValidationError(
            f"Required sheet {title!r} is missing; the workbook has "
            f"{', '.join(workbook.sheetnames)}.",
            field="sheets",
        )
    sheet = workbook[title]
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return SheetData(title=title, headers=(), rows=())
    headers = tuple(str(value).strip() for value in header_row if value is not None)
    width = len(headers)
    rows = [
        tuple(row[:width])
        for row in iterator
        if any(value is not None and str(value).strip() != "" for value in row[:width])
    ]
    return SheetData(title=title, headers=headers, rows=tuple(rows))


@contextmanager
def atomic_workbook_save(destination: Path) -> Iterator[Workbook]:
    """Build a workbook and replace ``destination`` with it in one step.

    Args:
        destination: Final path. Its parent directory is created if absent.

    Yields:
        A fresh workbook with its default sheet already removed, so the caller creates
        every sheet explicitly and sheet order is exactly what the caller wrote.
    """
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default = workbook.active
    if default is not None:
        workbook.remove(default)
    handle, temporary = tempfile.mkstemp(
        prefix=f".{destination.name}.", suffix=".partial", dir=destination.parent
    )
    os.close(handle)
    temporary_path = Path(temporary)
    try:
        yield workbook
        workbook.save(temporary_path)
        os.replace(temporary_path, destination)
    finally:
        workbook.close()
        if temporary_path.exists():
            temporary_path.unlink()


def write_title(sheet: Worksheet, text: str) -> None:
    """Write the sheet title into ``A1`` with the governed title style."""
    sheet["A1"] = text
    sheet["A1"].font = _TITLE_FONT


def write_heading(sheet: Worksheet, row: int, text: str) -> None:
    """Write a section heading into column A of ``row``."""
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = _HEADING_FONT


def write_notice(sheet: Worksheet, row: int, heading: str, text: str, *, width: int = 8) -> int:
    """Write the public-reference notice block and return the next free row.

    Args:
        sheet: Target sheet.
        row: First row of the block.
        heading: Bold heading above the notice.
        text: The notice itself.
        width: Number of columns the notice is merged across.

    Returns:
        The first row after the block.
    """
    write_heading(sheet, row, heading)
    cell = sheet.cell(row=row + 1, column=1, value=text)
    cell.font = _NOTICE_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(
        start_row=row + 1, start_column=1, end_row=row + 1, end_column=max(width, 2)
    )
    sheet.row_dimensions[row + 1].height = 46
    return row + 3


def style_key_value_block(sheet: Worksheet, first_row: int, last_row: int) -> None:
    """Apply the label style to column A of a key/value block."""
    for row in range(first_row, last_row + 1):
        sheet.cell(row=row, column=1).font = _LABEL_FONT


def style_table(  # noqa: PLR0913 - a table is placed by its four corners plus two flags
    sheet: Worksheet,
    *,
    name: str,
    first_row: int,
    last_row: int,
    last_column: int,
    first_column: int = 1,
    freeze: bool = True,
    autofilter: bool = True,
) -> None:
    """Apply the governed table style to a header row and its data rows.

    A real Excel table object is added -- not merely a fill -- so the produced workbook
    carries banded rows, a filter control and a named range a dealership user can pivot
    from. openpyxl refuses a table whose reference covers a single row, so a table with
    no data rows is styled without the table object rather than failing the export.

    Args:
        sheet: Target sheet.
        name: Excel table name; must be unique in the workbook and contain no spaces.
        first_row: Row holding the headers.
        last_row: Last data row.
        last_column: Last populated column, one-based.
        first_column: First populated column, one-based. A table's reference must start
            on its own header cell; a block that begins in column D and is declared from
            column A produces a workbook Excel refuses to open.
        freeze: Whether to freeze everything above the first data row.
        autofilter: Whether to add a filter control across the header row.
    """
    for column in range(first_column, last_column + 1):
        cell = sheet.cell(row=first_row, column=column)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in range(first_row, last_row + 1):
        for column in range(first_column, last_column + 1):
            sheet.cell(row=row, column=column).border = _CELL_BORDER

    start = f"{get_column_letter(first_column)}{first_row}"
    end = f"{get_column_letter(last_column)}{max(last_row, first_row + 1)}"
    reference = f"{start}:{end}"
    if last_row > first_row:
        table = Table(displayName=name, ref=reference)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
        )
        sheet.add_table(table)
    elif autofilter:
        sheet.auto_filter.ref = reference

    if freeze:
        sheet.freeze_panes = f"{get_column_letter(first_column)}{first_row + 1}"


def column_widths(sheet: Worksheet, widths: Sequence[int]) -> None:
    """Set controlled column widths from column A rightwards."""
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def add_condition_chart(
    sheet: Worksheet, *, anchor: str, labels: str, values: str, title: str
) -> None:
    """Add the condition-mix bar chart.

    Args:
        sheet: Sheet the chart is anchored on and whose cells it reads.
        anchor: Cell the chart's top-left corner sits on.
        labels: ``A1``-style range of the category labels.
        values: ``A1``-style range of the values, header row included.
        title: Chart title.
    """
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.height = 6.5
    chart.width = 11
    chart.legend = None
    chart.add_data(Reference(sheet, range_string=f"'{sheet.title}'!{values}"), titles_from_data=True)
    chart.set_categories(Reference(sheet, range_string=f"'{sheet.title}'!{labels}"))
    sheet.add_chart(chart, anchor)


def add_top_model_chart(
    sheet: Worksheet, *, anchor: str, labels: str, values: str, title: str
) -> None:
    """Add the top-model horizontal bar chart.

    Args:
        sheet: Sheet the chart is anchored on and whose cells it reads.
        anchor: Cell the chart's top-left corner sits on.
        labels: ``A1``-style range of the model labels.
        values: ``A1``-style range of the unit counts, header row included.
        title: Chart title.
    """
    chart = BarChart()
    chart.type = "bar"
    chart.title = title
    chart.height = 9
    chart.width = 13
    chart.legend = None
    chart.add_data(Reference(sheet, range_string=f"'{sheet.title}'!{values}"), titles_from_data=True)
    chart.set_categories(Reference(sheet, range_string=f"'{sheet.title}'!{labels}"))
    sheet.add_chart(chart, anchor)
