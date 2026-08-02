"""Private inventory workbook in, governed public-reference workbook out.

This is the only code in ARPI that ever touches an unsanitized dealership file, and it is
written on the assumption that it will be run on a laptop, in a hurry, against a file
nobody may commit.

What it removes, and what that costs
------------------------------------
==========================  =========================================================
Removed                     Replaced with
==========================  =========================================================
Original VIN                A deterministic, group-stable ``ARPI``-prefixed synthetic
                            VIN and vehicle identifier. No reverse mapping is produced
                            and no function here could build one.
Row-level source URL        A neutral feed label naming the lane, never the origin.
External dealer identity    The fictional Granite State Auto Group store the operator
                            assigns, resolved from the authoritative registry.
Street address              Nothing. Geography stops at store name and market region.
==========================  =========================================================

The cost is real and is stated on the artifact: the result cannot be traced back to a
source listing, which is the point, and therefore cannot be re-verified against one.

Order of operations
-------------------
1. Validate the private input **before transforming anything**, so a malformed file fails
   with a redacted message rather than half-writing a sanitized one.
2. Resolve the store from the ARPI registry, never from the input or the file name.
3. Derive identity and batch metadata deterministically.
4. Build the four-sheet governed workbook in memory.
5. Write atomically, refusing to overwrite unless told to.

Nothing is printed
------------------
No function here returns, logs or formats an original VIN, a source URL or a complete
source row. Errors name a row number, a column and a category. :class:`SanitizationSummary`
is the whole of what a run reports, and it holds counts and repository paths.
"""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl.styles import Alignment

from arpi.constants import CHECK_CATEGORY_COMPLETENESS, CHECK_CATEGORY_STRUCTURAL
from arpi.exceptions import ValidationError
from arpi.inventory.contract import InventoryListingContract, load_contract, normalise_header
from arpi.inventory.identity import (
    default_output_path,
    derived_sanitized_file_name,
    source_batch_id,
    source_record_id,
    synthetic_identity,
)
from arpi.inventory.spec import resolve_store
from arpi.inventory.validation import (
    ValidationFinding,
    contains_url,
    file_digest,
    looks_like_a_real_vin,
)
from arpi.inventory.workbook import (
    CURRENCY_FORMAT,
    DATE_FORMAT,
    INTEGER_FORMAT,
    atomic_workbook_save,
    column_widths,
    open_read_only,
    read_sheet_rows,
    style_key_value_block,
    style_table,
    write_heading,
    write_notice,
    write_title,
)

__all__ = ["SanitizationSummary", "sanitize_workbook"]

#: The notice every governed workbook carries, verbatim, on its README sheet.
PUBLIC_REFERENCE_NOTICE = (
    "Granite State Auto Group and its stores are fictional. Original VINs, source URLs, "
    "street address and real dealership identity have been removed. This workbook is NOT "
    "fully synthetic: model, trim, mileage, condition, advertised-price and inventory-mix "
    "attributes are retained from a de-identified public reference snapshot. Do not "
    "present these rows as current dealership performance, confidential operational data, "
    "completed sales, or transaction results."
)

#: The business-semantics block, stated on the artifact rather than only in a document.
_SEMANTICS: tuple[tuple[str, str], ...] = (
    (
        "Observed listing",
        "A row proves that a vehicle listing was visible in the source snapshot. It does "
        "not prove physical on-ground status and does not prove dealership ownership.",
    ),
    (
        "Advertised price",
        "The listed price is not transaction price, acquisition cost, inventory "
        "investment, MSRP, or gross.",
    ),
    (
        "Removed listing",
        "A future disappearance means REMOVED FROM LISTING, not sold. It could reflect a "
        "sale, a trade, wholesale, feed suppression, or an error.",
    ),
    (
        "Days observed online",
        "Repeated snapshots can calculate days observed online. That must not be labelled "
        "days in stock.",
    ),
    (
        "Portfolio status",
        "Dealer and vehicle identifiers are synthetic. Model, trim, mileage, condition and "
        "advertised-price attributes remain a de-identified public reference snapshot.",
    ),
)


@dataclass(frozen=True, slots=True)
class SanitizationSummary:
    """What one sanitizer run did, with nothing in it that could identify a source.

    Attributes:
        dealership_id: Store the rows were assigned to.
        store_name: Store name resolved from the registry.
        captured_at: Snapshot date.
        source_batch_id: Deterministic batch identifier stamped on every row.
        row_count: Rows read from the private input and written to the artifact.
        identifiers_replaced: Number of original vehicle identifiers replaced.
        urls_removed: Number of row-level source URLs removed.
        output_path: Where the artifact was written, or would be written on a dry run.
        output_file_name: The approved underscore-based file name.
        sheets: Sheet names written, in order.
        digest: SHA-256 of the written bytes, or ``None`` on a dry run.
        dry_run: Whether anything was written at all.
    """

    dealership_id: str
    store_name: str
    captured_at: date
    source_batch_id: str
    row_count: int
    identifiers_replaced: int
    urls_removed: int
    output_path: Path
    output_file_name: str
    sheets: tuple[str, ...]
    digest: str | None
    dry_run: bool

    def summary(self) -> str:
        """Render a redacted, human-readable summary."""
        lines = [
            f"mode               : {'dry run (nothing written)' if self.dry_run else 'write'}",
            f"dealership         : {self.dealership_id} ({self.store_name})",
            f"captured at        : {self.captured_at.isoformat()}",
            f"source batch       : {self.source_batch_id}",
            f"rows               : {self.row_count}",
            f"identifiers replaced: {self.identifiers_replaced}",
            f"urls removed       : {self.urls_removed}",
            f"sheets             : {', '.join(self.sheets)}",
            f"output file name   : {self.output_file_name}",
            f"output path        : {self.output_path.as_posix()}",
        ]
        if self.digest:
            lines.append(f"sha256             : {self.digest}")
        return "\n".join(lines)

    def as_dict(self) -> dict[str, Any]:
        """Render the summary as a JSON-serialisable mapping."""
        return {
            "mode": "dry-run" if self.dry_run else "write",
            "dealership_id": self.dealership_id,
            "store_name": self.store_name,
            "captured_at": self.captured_at.isoformat(),
            "source_batch_id": self.source_batch_id,
            "rows": self.row_count,
            "identifiers_replaced": self.identifiers_replaced,
            "urls_removed": self.urls_removed,
            "sheets": list(self.sheets),
            "output_file_name": self.output_file_name,
            "output_path": self.output_path.as_posix(),
            "sha256": self.digest,
        }


# --------------------------------------------------------------------------------------
# Private input
# --------------------------------------------------------------------------------------


def _private_header_index(
    headers: Sequence[str], contract: InventoryListingContract
) -> dict[str, int]:
    """Map each declared private role to its column position, refusing a missing one."""
    normalised = {normalise_header(header): index for index, header in enumerate(headers)}
    index: dict[str, int] = {}
    missing: list[str] = []
    for column in contract.private_input_columns:
        position = normalised.get(normalise_header(column.header))
        if position is None:
            if column.required:
                missing.append(column.header)
            continue
        index[column.role] = position
    if missing:
        raise ValidationError(
            f"The private workbook is missing {len(missing)} required column(s): "
            f"{', '.join(missing)}. The accepted input contract is "
            f"{', '.join(column.header for column in contract.private_input_columns)}.",
            field="private_input",
        )
    return index


def _text(value: Any) -> str:
    """Render a cell value as trimmed text."""
    return "" if value is None else str(value).strip()


def _int_or_none(value: Any) -> int | None:
    """Coerce to int, tolerating ``"12,345"`` and ``"12345 mi"``-style noise."""
    text = _text(value).replace(",", "")
    if not text:
        return None
    cleaned = "".join(character for character in text if character.isdigit() or character == "-")
    if not cleaned or cleaned == "-":
        return None
    try:
        return int(cleaned)
    except ValueError:
        return None


def _price_or_none(value: Any) -> float | None:
    """Coerce to a price, tolerating ``"$38,690"`` and returning ``None`` for prose."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value).replace(",", "").replace("$", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _capture_date(value: Any, fallback: date) -> date:
    """Read a per-row capture date, falling back to the operator-supplied one."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return fallback
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return fallback


@dataclass(frozen=True, slots=True)
class _SanitizedRow:
    """One transformed row, ready to be written to the Inventory sheet."""

    values: tuple[Any, ...]
    had_url: bool


def _transform(  # noqa: PLR0913 - one argument per contract concern, all required
    rows: Sequence[tuple[Any, ...]],
    *,
    index: dict[str, int],
    dealership_id: str,
    store_name: str,
    captured_at: date,
    batch_id: str,
    contract: InventoryListingContract,
) -> tuple[tuple[_SanitizedRow, ...], tuple[ValidationFinding, ...]]:
    """Transform private rows into sanitized rows, collecting redacted findings."""
    findings: list[ValidationFinding] = []
    sanitized: list[_SanitizedRow] = []
    seen_identities: dict[str, int] = {}

    def fail(row_number: int, column: str, category: str, message: str) -> None:
        findings.append(
            ValidationFinding(
                check_id="CONTRACT",
                category=category,
                message=message,
                sheet="private input",
                row=row_number,
                column=column,
            )
        )

    for row_number, row in enumerate(rows, start=1):

        def cell(role: str, source_row: tuple[Any, ...] = row) -> Any:
            position = index.get(role)
            return None if position is None else source_row[position]

        original_vin = _text(cell("original_vin"))
        if not original_vin:
            fail(row_number, "VIN", CHECK_CATEGORY_COMPLETENESS, "the source identifier is absent")
            continue
        if not looks_like_a_real_vin(original_vin) and len(original_vin) != 17:
            # Not a refusal on its own -- powersports and older units carry shorter
            # identifiers -- but a zero-length or obviously truncated value is.
            if len(original_vin) < 8:
                fail(
                    row_number,
                    "VIN",
                    CHECK_CATEGORY_STRUCTURAL,
                    "the source identifier is too short to be a vehicle identifier",
                )
                continue

        identity = synthetic_identity(original_vin, contract=contract)
        if identity.vehicle_id in seen_identities:
            fail(
                row_number,
                "VIN",
                CHECK_CATEGORY_STRUCTURAL,
                "the source identifier duplicates the one first seen on row "
                f"{seen_identities[identity.vehicle_id]}; one physical vehicle cannot be "
                "listed twice in one snapshot",
            )
            continue
        seen_identities[identity.vehicle_id] = row_number

        condition = _text(cell("condition")).title()
        if condition not in contract.condition_values:
            fail(
                row_number,
                "Condition",
                CHECK_CATEGORY_STRUCTURAL,
                f"Condition is outside the governed domain "
                f"({' | '.join(contract.condition_values)})",
            )
            continue

        model_year = _int_or_none(cell("model_year"))
        if model_year is None:
            fail(row_number, "Year", CHECK_CATEGORY_STRUCTURAL, "Year is not an integer")
            continue

        make = _text(cell("make"))
        model = _text(cell("model"))
        if not make:
            fail(row_number, "Make", CHECK_CATEGORY_COMPLETENESS, "Make is blank")
            continue
        if not model:
            fail(row_number, "Model", CHECK_CATEGORY_COMPLETENESS, "Model is blank")
            continue

        trim = _text(cell("trim"))
        display = _text(cell("vehicle_display")) or " ".join(
            part for part in (str(model_year), make, model, trim) if part
        )

        odometer = _int_or_none(cell("odometer_miles"))
        if odometer is None or odometer < contract.odometer_minimum:
            fail(
                row_number,
                "Mileage",
                CHECK_CATEGORY_STRUCTURAL,
                "Mileage is absent or negative",
            )
            continue

        pricing_status = _text(cell("pricing_status")) or contract.pricing_status_values[0]
        matched = [
            value
            for value in contract.pricing_status_values
            if value.casefold() == pricing_status.casefold()
        ]
        if not matched:
            fail(
                row_number,
                "Price Status",
                CHECK_CATEGORY_STRUCTURAL,
                f"Price Status is outside the governed domain "
                f"({' | '.join(contract.pricing_status_values)})",
            )
            continue
        pricing_status = matched[0]
        listed = pricing_status == contract.pricing_status_values[0]

        price = _price_or_none(cell("advertised_price"))
        if listed and price is None:
            fail(
                row_number,
                "Price",
                CHECK_CATEGORY_COMPLETENESS,
                "Price Status is 'Listed' but no price is present",
            )
            continue
        if not listed:
            # The contract does not permit a priced call-for-price row. Dropping the
            # number is the sanitization, not a rejection: the two columns disagreed in
            # the source and the status is the governing statement.
            price = None

        had_url = contains_url(cell("source_url"))
        row_capture = _capture_date(cell("captured_at"), captured_at)
        if row_capture != captured_at:
            fail(
                row_number,
                "Captured",
                CHECK_CATEGORY_STRUCTURAL,
                "the row's capture date disagrees with the snapshot date supplied to the "
                "sanitizer; one workbook is one snapshot",
            )
            continue

        sanitized.append(
            _SanitizedRow(
                values=(
                    source_record_id(dealership_id, captured_at, len(sanitized) + 1),
                    dealership_id,
                    store_name,
                    captured_at,
                    batch_id,
                    contract.source_feed,
                    condition,
                    model_year,
                    make,
                    model,
                    trim or None,
                    display,
                    odometer,
                    price,
                    pricing_status,
                    identity.vehicle_id,
                    identity.vin,
                    contract.inventory_unit_count,
                    contract.classification,
                ),
                had_url=had_url,
            )
        )

    return tuple(sanitized), tuple(findings)


# --------------------------------------------------------------------------------------
# Workbook construction
# --------------------------------------------------------------------------------------


def _write_readme(  # noqa: PLR0913 - the README block is one statement with many fields
    sheet: Any,
    *,
    dealership_id: str,
    store_name: str,
    market_region: str,
    captured_at: date,
    row_count: int,
    urls_removed: int,
    output_path: Path,
    contract: InventoryListingContract,
) -> None:
    """Write the governance sheet: provenance, controls applied, semantics, notice."""
    write_title(sheet, "ARPI Inventory Operations | Sanitized Public Reference Artifact")
    facts = (
        ("Artifact", f"{store_name} inventory listing snapshot"),
        ("Dealer group", "Granite State Auto Group"),
        ("Store", store_name),
        ("Dealership ID", dealership_id),
        ("Market", market_region),
        ("Snapshot date", captured_at),
        ("Source type", "Public inventory listing snapshot, de-identified for portfolio use"),
        ("Classification", contract.classification),
        ("Sanitization version", contract.contract_version),
    )
    for offset, (label, value) in enumerate(facts):
        sheet.cell(row=4 + offset, column=1, value=label)
        cell = sheet.cell(row=4 + offset, column=2, value=value)
        if isinstance(value, date):
            cell.number_format = DATE_FORMAT
    style_key_value_block(sheet, 4, 3 + len(facts))

    row = 5 + len(facts)
    write_heading(sheet, row, "Sanitization controls applied")
    controls = (
        ("Control", "Result", "Public artifact treatment"),
        (
            "Real VINs",
            f"{row_count} replaced",
            "Replaced with deterministic ARPI-prefixed synthetic VINs; no original VIN "
            "retained and no reverse mapping produced.",
        ),
        (
            "Source URLs",
            f"{urls_removed} removed",
            "Row-level external URLs removed and replaced with a neutral source-feed label.",
        ),
        (
            "External dealer identity",
            "Removed",
            f"All external dealership references removed. Records assigned to the "
            f"fictional store {dealership_id}.",
        ),
        ("Street address", "Removed", "Geography stops at store name and market region."),
        (
            "Record identity",
            f"{row_count} generated",
            "Synthetic source record IDs and group-stable vehicle IDs support repeatable "
            "ingestion, cross-snapshot matching and future multi-store analysis.",
        ),
    )
    for offset, control in enumerate(controls):
        for column, value in enumerate(control, start=1):
            sheet.cell(row=row + 1 + offset, column=column, value=value)
    style_table(
        sheet,
        name="SanitizationControls",
        first_row=row + 1,
        last_row=row + len(controls),
        last_column=3,
        freeze=False,
        autofilter=False,
    )

    row = row + len(controls) + 3
    write_heading(sheet, row, "Business semantics and limitations")
    for offset, (label, text) in enumerate(_SEMANTICS):
        sheet.cell(row=row + 1 + offset, column=1, value=label)
        sheet.cell(row=row + 1 + offset, column=2, value=text)
    style_key_value_block(sheet, row + 1, row + len(_SEMANTICS))

    row = row + len(_SEMANTICS) + 3
    write_heading(sheet, row, "Repository placement and intended ingestion contract")
    placement = (
        ("Recommended repository path", output_path.as_posix()),
        ("Canonical file name", output_path.name),
        (
            "Naming convention",
            "ARPI_<Store_Descriptor>_Inventory_Sanitized_<yyyy-mm-dd>.xlsx -- underscores "
            "between filename words, hyphens only inside the ISO date.",
        ),
        (
            "Raw source policy",
            "Keep the original unsanitized workbook outside Git and outside the repository.",
        ),
        ("Intended adapter", "scripts/import_inventory_snapshot.py"),
        ("Target source entity", "inventory_listing_snapshot"),
        ("Target warehouse fact", "warehouse.fact_vehicle_listing_snapshot"),
        (
            "Governance",
            "ADR-0011 governs this lane and data/reference/README.md is its policy. This "
            "artifact must not be placed under data/sample, which is reserved for fully "
            "machine-generated data.",
        ),
    )
    for offset, (label, value) in enumerate(placement):
        sheet.cell(row=row + 1 + offset, column=1, value=label)
        sheet.cell(row=row + 1 + offset, column=2, value=value)
    style_key_value_block(sheet, row + 1, row + len(placement))

    write_notice(sheet, row + len(placement) + 2, "Published reference notice", PUBLIC_REFERENCE_NOTICE)
    column_widths(sheet, (30, 62, 62))


def _write_inventory(
    sheet: Any, rows: Sequence[_SanitizedRow], contract: InventoryListingContract
) -> None:
    """Write the Inventory sheet: exact headers, typed cells, styled table."""
    for column, header in enumerate(contract.headers, start=1):
        sheet.cell(row=1, column=column, value=header)
    for offset, sanitized in enumerate(rows):
        for column, value in enumerate(sanitized.values, start=1):
            cell = sheet.cell(row=2 + offset, column=column, value=value)
            declared = contract.inventory_columns[column - 1]
            if declared.type_name == "date":
                cell.number_format = DATE_FORMAT
            elif declared.type_name == "money":
                cell.number_format = CURRENCY_FORMAT
            elif declared.type_name == "integer":
                cell.number_format = INTEGER_FORMAT
    style_table(
        sheet,
        name="InventoryListings",
        first_row=1,
        last_row=1 + len(rows),
        last_column=len(contract.headers),
    )
    column_widths(
        sheet,
        (22, 14, 30, 13, 22, 34, 11, 11, 18, 30, 18, 42, 15, 17, 15, 20, 20, 19, 30),
    )


def _write_summary(  # noqa: PLR0913 - the summary block is one statement with many fields
    sheet: Any,
    *,
    store_name: str,
    dealership_id: str,
    market_region: str,
    captured_at: date,
    row_count: int,
    top_models: Sequence[tuple[str, str, str]],
    contract: InventoryListingContract,
) -> None:
    """Write the formula-driven Summary sheet."""
    last = row_count + 1
    inventory = contract.sheets["inventory"]
    condition_range = f"'{inventory}'!$G$2:$G${last}"
    status_range = f"'{inventory}'!$O$2:$O${last}"
    price_range = f"'{inventory}'!$N$2:$N${last}"
    id_range = f"'{inventory}'!$A$2:$A${last}"

    write_title(sheet, f"ARPI Inventory Operations | {store_name}")
    facts = (
        ("Dealer group", "Granite State Auto Group"),
        ("Store", store_name),
        ("Dealership ID", dealership_id),
        ("Market", market_region),
        ("Snapshot date", captured_at),
    )
    for offset, (label, value) in enumerate(facts):
        sheet.cell(row=4 + offset, column=1, value=label)
        cell = sheet.cell(row=4 + offset, column=2, value=value)
        if isinstance(value, date):
            cell.number_format = DATE_FORMAT
    style_key_value_block(sheet, 4, 3 + len(facts))

    sheet.cell(row=4, column=4, value="Condition")
    sheet.cell(row=4, column=5, value="Count")
    condition_rows = (
        ("New", f'=COUNTIF({condition_range},"New")'),
        ("Used", f'=COUNTIF({condition_range},"Used")'),
        ("Listed price", f'=COUNTIF({status_range},"Listed")'),
        ("Call for price", f'=COUNTIF({status_range},"Call for price")'),
    )
    for offset, (label, formula) in enumerate(condition_rows):
        sheet.cell(row=5 + offset, column=4, value=label)
        sheet.cell(row=5 + offset, column=5, value=formula).number_format = INTEGER_FORMAT
    style_table(
        sheet,
        name="ConditionMix",
        first_row=4,
        last_row=4 + len(condition_rows),
        first_column=4,
        last_column=5,
        freeze=False,
        autofilter=False,
    )

    sheet.cell(row=10, column=1, value="Metric")
    sheet.cell(row=10, column=2, value="Value")
    metrics = (
        ("Total listed vehicles", f"=COUNTA({id_range})", INTEGER_FORMAT),
        ("New vehicles", f'=COUNTIF({condition_range},"New")', INTEGER_FORMAT),
        ("Used vehicles", f'=COUNTIF({condition_range},"Used")', INTEGER_FORMAT),
        ("Vehicles with listed price", f'=COUNTIF({status_range},"Listed")', INTEGER_FORMAT),
        ("Call for price", f'=COUNTIF({status_range},"Call for price")', INTEGER_FORMAT),
        ("Total advertised listing value", f"=SUM({price_range})", CURRENCY_FORMAT),
        ("Average advertised price", "=IF(B14=0,0,B16/B14)", CURRENCY_FORMAT),
        ("Pricing completeness", "=IF(B11=0,0,B14/B11)", "0.0%"),
    )
    for offset, (label, formula, number_format) in enumerate(metrics):
        sheet.cell(row=11 + offset, column=1, value=label)
        sheet.cell(row=11 + offset, column=2, value=formula).number_format = number_format
    style_table(
        sheet,
        name="SummaryMetrics",
        first_row=10,
        last_row=10 + len(metrics),
        last_column=2,
        freeze=False,
        autofilter=False,
    )

    row = 11 + len(metrics) + 1
    write_heading(sheet, row, "Governed interpretation")
    cell = sheet.cell(
        row=row + 1,
        column=1,
        value=(
            f"This report describes an observed public listing snapshot for the fictional "
            f"{store_name}. Advertised prices are not transaction prices or inventory cost. "
            "Listing disappearance does not prove a sale. Repeated snapshots may support "
            "listing churn and days observed online, but not dealership days in stock."
        ),
    )
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=6)
    sheet.row_dimensions[row + 1].height = 46

    row = row + 4
    write_heading(sheet, row, "Top inventory models")
    sheet.cell(row=row + 1, column=1, value="Condition")
    sheet.cell(row=row + 1, column=2, value="Make")
    sheet.cell(row=row + 1, column=3, value="Model")
    sheet.cell(row=row + 1, column=4, value="Units")
    make_range = f"'{inventory}'!$I$2:$I${last}"
    model_range = f"'{inventory}'!$J$2:$J${last}"
    for offset, (condition, make, model) in enumerate(top_models):
        target = row + 2 + offset
        sheet.cell(row=target, column=1, value=condition)
        sheet.cell(row=target, column=2, value=make)
        sheet.cell(row=target, column=3, value=model)
        sheet.cell(
            row=target,
            column=4,
            value=(
                f"=COUNTIFS({condition_range},A{target},{make_range},B{target},"
                f"{model_range},C{target})"
            ),
        ).number_format = INTEGER_FORMAT
    style_table(
        sheet,
        name="TopModels",
        first_row=row + 1,
        last_row=row + 1 + len(top_models),
        last_column=4,
        freeze=False,
        autofilter=False,
    )

    write_notice(
        sheet,
        row + len(top_models) + 4,
        "Sanitized public reference notice",
        PUBLIC_REFERENCE_NOTICE,
    )
    column_widths(sheet, (32, 26, 32, 14, 16, 18))


def _write_model_summary(
    sheet: Any,
    rows: Sequence[_SanitizedRow],
    contract: InventoryListingContract,
) -> None:
    """Write the formula-driven Model Summary sheet, one row per condition/make/model."""
    last = len(rows) + 1
    inventory = contract.sheets["inventory"]
    condition_range = f"'{inventory}'!$G$2:$G${last}"
    make_range = f"'{inventory}'!$I$2:$I${last}"
    model_range = f"'{inventory}'!$J$2:$J${last}"
    status_range = f"'{inventory}'!$O$2:$O${last}"
    price_range = f"'{inventory}'!$N$2:$N${last}"

    write_title(sheet, "ARPI Inventory Operations | Model Summary")
    headers = (
        "Condition",
        "Make",
        "Model",
        "Total Units",
        "Listed Units",
        "Call for Price",
        "Average Listed Price",
        "Minimum Listed Price",
        "Maximum Listed Price",
    )
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=4, column=column, value=header)

    groups = _model_groups(rows)
    for offset, (condition, make, model) in enumerate(groups):
        target = 5 + offset
        sheet.cell(row=target, column=1, value=condition)
        sheet.cell(row=target, column=2, value=make)
        sheet.cell(row=target, column=3, value=model)
        criteria = (
            f"{condition_range},$A{target},{make_range},$B{target},{model_range},$C{target}"
        )
        sheet.cell(row=target, column=4, value=f"=COUNTIFS({criteria})").number_format = (
            INTEGER_FORMAT
        )
        sheet.cell(
            row=target, column=5, value=f'=COUNTIFS({criteria},{status_range},"Listed")'
        ).number_format = INTEGER_FORMAT
        sheet.cell(
            row=target, column=6, value=f'=COUNTIFS({criteria},{status_range},"Call for price")'
        ).number_format = INTEGER_FORMAT
        sheet.cell(
            row=target,
            column=7,
            value=(
                f'=IF(E{target}=0,"",SUMIFS({price_range},{criteria},{status_range},'
                f'"Listed")/E{target})'
            ),
        ).number_format = CURRENCY_FORMAT
        sheet.cell(
            row=target,
            column=8,
            value=f'=IF(E{target}=0,"",MINIFS({price_range},{criteria},{status_range},"Listed"))',
        ).number_format = CURRENCY_FORMAT
        sheet.cell(
            row=target,
            column=9,
            value=f'=IF(E{target}=0,"",MAXIFS({price_range},{criteria},{status_range},"Listed"))',
        ).number_format = CURRENCY_FORMAT

    style_table(
        sheet,
        name="InventoryModelSummary",
        first_row=4,
        last_row=4 + len(groups),
        last_column=len(headers),
    )
    column_widths(sheet, (12, 20, 34, 13, 13, 15, 20, 20, 20))


def _model_groups(rows: Sequence[_SanitizedRow]) -> tuple[tuple[str, str, str], ...]:
    """Group rows by condition, make and model, ordered by unit count then name.

    Deterministic ordering matters: the exporter and the sanitizer must produce the same
    sheet for the same data, and a set-iteration order would make that untestable.
    """
    counts: dict[tuple[str, str, str], int] = {}
    for row in rows:
        key = (str(row.values[6]), str(row.values[8]), str(row.values[9]))
        counts[key] = counts.get(key, 0) + 1
    return tuple(
        key
        for key in sorted(
            counts,
            key=lambda item: (0 if item[0] == "New" else 1, -counts[item], item[1], item[2]),
        )
    )


# --------------------------------------------------------------------------------------
# Entry point
# --------------------------------------------------------------------------------------


def sanitize_workbook(  # noqa: PLR0913 - every argument is an operator decision
    *,
    input_path: Path,
    dealership_id: str,
    captured_at: date,
    output_path: Path | None = None,
    contract: InventoryListingContract | None = None,
    overwrite: bool = False,
    dry_run: bool = False,
    sheet_name: str | None = None,
) -> SanitizationSummary:
    """Sanitize one private inventory workbook into a governed public-reference artifact.

    Args:
        input_path: The private workbook. It stays outside the repository.
        dealership_id: Fictional store the rows are assigned to, e.g. ``"GSA-002"``.
        captured_at: Snapshot date.
        output_path: Where to write. Defaults to the governed path and the approved
            underscore-based file name derived from the store and the date.
        contract: Contract to write against.
        overwrite: Whether an existing output may be replaced. Default is refusal.
        dry_run: Report what would happen and write nothing.
        sheet_name: Sheet of the private workbook to read. Defaults to the first sheet.

    Returns:
        A redacted summary of the run.

    Raises:
        ValidationError: If the private input is malformed, if any row fails the input
            contract, or if the output exists and ``overwrite`` is not set. Messages name
            a row, a column and a category, never a value.
    """
    active = contract or load_contract()
    store = resolve_store(dealership_id)
    destination = output_path or default_output_path(
        store.dealership_id, captured_at, contract=active
    )
    expected_name = derived_sanitized_file_name(
        store.dealership_id, captured_at, contract=active
    )

    if destination.exists() and not overwrite and not dry_run:
        raise ValidationError(
            f"{destination.as_posix()} already exists. A sanitized reference artifact is "
            "never silently replaced: rerun with --overwrite if you intend to supersede "
            "it, and record why in the pull request.",
            field="output",
        )

    with open_read_only(input_path) as workbook:
        title = sheet_name or workbook.sheetnames[0]
        sheet = read_sheet_rows(workbook, title)

    index = _private_header_index(sheet.headers, active)
    batch_id = source_batch_id(store.dealership_id, captured_at)
    sanitized, findings = _transform(
        sheet.rows,
        index=index,
        dealership_id=store.dealership_id,
        store_name=store.store_name,
        captured_at=captured_at,
        batch_id=batch_id,
        contract=active,
    )
    if findings:
        rendered = "\n  ".join(finding.render() for finding in findings)
        raise ValidationError(
            f"The private workbook failed the input contract; {len(findings)} row(s) "
            f"could not be sanitized:\n  {rendered}",
            field="private_input",
        )
    if not sanitized:
        raise ValidationError(
            f"{input_path.name} produced no sanitizable rows. A sanitized artifact with "
            "no listings would be indistinguishable from a store with no inventory.",
            field="private_input",
        )

    urls_removed = sum(1 for row in sanitized if row.had_url)
    sheets = active.required_sheets

    if dry_run:
        return SanitizationSummary(
            dealership_id=store.dealership_id,
            store_name=store.store_name,
            captured_at=captured_at,
            source_batch_id=batch_id,
            row_count=len(sanitized),
            identifiers_replaced=len(sanitized),
            urls_removed=urls_removed,
            output_path=destination,
            output_file_name=expected_name,
            sheets=sheets,
            digest=None,
            dry_run=True,
        )

    top_models = _model_groups(sanitized)[:10]
    with atomic_workbook_save(destination) as book:
        _write_readme(
            book.create_sheet(active.sheets["readme"]),
            dealership_id=store.dealership_id,
            store_name=store.store_name,
            market_region=store.market_region,
            captured_at=captured_at,
            row_count=len(sanitized),
            urls_removed=urls_removed,
            output_path=destination,
            contract=active,
        )
        _write_summary(
            book.create_sheet(active.sheets["summary"]),
            store_name=store.store_name,
            dealership_id=store.dealership_id,
            market_region=store.market_region,
            captured_at=captured_at,
            row_count=len(sanitized),
            top_models=top_models,
            contract=active,
        )
        _write_inventory(book.create_sheet(active.sheets["inventory"]), sanitized, active)
        _write_model_summary(book.create_sheet(active.sheets["model_summary"]), sanitized, active)

    return SanitizationSummary(
        dealership_id=store.dealership_id,
        store_name=store.store_name,
        captured_at=captured_at,
        source_batch_id=batch_id,
        row_count=len(sanitized),
        identifiers_replaced=len(sanitized),
        urls_removed=urls_removed,
        output_path=destination,
        output_file_name=expected_name,
        sheets=sheets,
        digest=file_digest(destination),
        dry_run=False,
    )
