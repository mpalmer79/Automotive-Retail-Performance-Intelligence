"""The database-backed Excel operating report a dealership can actually use.

Everything on the produced workbook comes from ``reporting.*`` views over
``warehouse.fact_vehicle_listing_snapshot``. Nothing is recomputed in Python, and nothing
is read from the sanitized input workbook: the report is evidence that the warehouse holds
what the workbook carried, and a report assembled from the workbook would prove nothing.

Why the output mirrors the sanitized artifact
---------------------------------------------
The governed sanitized workbook is the shape an operator has already learned to read --
README, Summary, Inventory, Model Summary, in that order, with the same headers and the
same notice. The exporter reproduces that contract rather than inventing a second layout,
and adds one sheet the input cannot have: ``Snapshot Changes``, which exists only when a
prior capture is there to compare against.

Formula-driven where formulas belong
------------------------------------
The Summary and Model Summary sheets are formulas over the Inventory sheet, exactly as in
the sanitized artifact. That is not decoration: a dealership user filters the Inventory
table and expects the totals to move. A workbook of hard-coded numbers looks identical
until somebody edits it, at which point it is silently wrong.

The Snapshot Changes sheet is values, not formulas, because it compares two captures and
Excel has nothing to compare against -- the prior capture is not in the workbook.

What never reaches the file
---------------------------
No original VIN, no source URL, no connection string, no credential, no hidden source
identity, and no claim of a sale. The last one is structural: the change sheet's labels
come from ``reporting.vw_vehicle_listing_change``, which has no sold label to emit.
"""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl.styles import Alignment

from arpi.exceptions import ValidationError
from arpi.inventory.contract import InventoryListingContract, load_contract
from arpi.inventory.identity import derived_report_file_name
from arpi.inventory.sanitizer import PUBLIC_REFERENCE_NOTICE
from arpi.inventory.spec import resolve_store
from arpi.inventory.workbook import (
    CURRENCY_FORMAT,
    DATE_FORMAT,
    INTEGER_FORMAT,
    PERCENT_FORMAT,
    add_condition_chart,
    add_top_model_chart,
    atomic_workbook_save,
    column_widths,
    style_key_value_block,
    style_table,
    write_heading,
    write_notice,
    write_title,
)

__all__ = ["ReportSummary", "export_operating_report"]

#: One-based column positions on the Inventory sheet that carry a formatted type.
#: Named because a bare `column == 8` in a formatting loop says nothing about which
#: column moved when the sheet's shape changes.
_INVENTORY_INTEGER_COLUMNS = frozenset({7, 12})
_INVENTORY_CURRENCY_COLUMN = 8
_INVENTORY_DATE_COLUMNS = frozenset({13, 14})

#: The same, for the Snapshot Changes sheet.
_CHANGE_CURRENCY_COLUMNS = frozenset({8, 9, 10})
_CHANGE_DATE_COLUMN = 13
_CHANGE_INTEGER_COLUMN = 14

#: The interpretation block the Summary sheet carries, stated on the artifact rather than
#: only in a document a reader of the workbook will never open.
GOVERNED_INTERPRETATION = (
    "This report describes an OBSERVED PUBLIC LISTING SNAPSHOT for a fictional store. "
    "Advertised prices are not transaction prices, acquisition cost, inventory investment "
    "or gross. A listing that disappears was REMOVED FROM LISTING, which is not the same as "
    "sold. Days observed online is not days in stock. Nothing here establishes current "
    "business performance."
)


@dataclass(frozen=True, slots=True)
class ReportSummary:
    """What one export produced.

    Attributes:
        dealership_id: Store the report covers.
        store_name: Store name resolved from the registry.
        captured_at: Capture date the report is as of.
        prior_captured_at: The preceding capture, when one exists.
        row_count: Listing rows written to the Inventory sheet.
        model_group_count: Rows written to the Model Summary sheet.
        change_row_count: Rows written to the Snapshot Changes sheet, or zero.
        sheets: Sheet names written, in order.
        output_path: Where the workbook was written.
        output_file_name: The approved underscore-based file name.
    """

    dealership_id: str
    store_name: str
    captured_at: date
    prior_captured_at: date | None
    row_count: int
    model_group_count: int
    change_row_count: int
    sheets: tuple[str, ...]
    output_path: Path
    output_file_name: str

    def summary(self) -> str:
        """Render a redacted, human-readable summary."""
        return "\n".join(
            [
                f"dealership         : {self.dealership_id} ({self.store_name})",
                f"captured at        : {self.captured_at.isoformat()}",
                f"prior capture      : "
                f"{self.prior_captured_at.isoformat() if self.prior_captured_at else 'none'}",
                f"listing rows       : {self.row_count}",
                f"model groups       : {self.model_group_count}",
                f"change rows        : {self.change_row_count}",
                f"sheets             : {', '.join(self.sheets)}",
                f"output file name   : {self.output_file_name}",
                f"output path        : {self.output_path.as_posix()}",
            ]
        )

    def as_dict(self) -> dict[str, Any]:
        """Render the summary as a JSON-serialisable mapping."""
        return {
            "dealership_id": self.dealership_id,
            "store_name": self.store_name,
            "captured_at": self.captured_at.isoformat(),
            "prior_captured_at": (
                self.prior_captured_at.isoformat() if self.prior_captured_at else None
            ),
            "listing_rows": self.row_count,
            "model_groups": self.model_group_count,
            "change_rows": self.change_row_count,
            "sheets": list(self.sheets),
            "output_file_name": self.output_file_name,
            "output_path": self.output_path.as_posix(),
        }


# --------------------------------------------------------------------------------------
# Queries. Every one reads reporting.*, never warehouse.* and never the input workbook.
# --------------------------------------------------------------------------------------

_LISTING_SQL = """
    SELECT c.condition_type, c.model_year, c.make, c.model, c.trim, c.vehicle_display,
           c.odometer_miles, c.advertised_price, c.pricing_status,
           c.synthetic_vehicle_id, c.synthetic_vin, c.inventory_unit_count,
           c.first_observed_at, c.last_observed_at, c.source_batch_id, c.source_file_name
    FROM reporting.vw_vehicle_listing_current AS c
    JOIN reporting.vw_dealership AS d ON d.dealership_key = c.dealership_key
    WHERE d.dealership_code = %s AND c.last_captured_at = %s
    ORDER BY c.condition_type, c.make, c.model, c.trim NULLS FIRST, c.synthetic_vehicle_id
"""

_SUMMARY_SQL = """
    SELECT s.observed_listing_units, s.new_listing_units, s.used_listing_units,
           s.listed_price_units, s.call_for_price_units, s.price_not_exposed_units,
           s.unpriced_units, s.total_advertised_value,
           s.minimum_advertised_price, s.maximum_advertised_price,
           s.latest_capture_age_days, s.is_latest_snapshot, s.source_file_name
    FROM reporting.vw_vehicle_listing_summary AS s
    JOIN reporting.vw_dealership AS d ON d.dealership_key = s.dealership_key
    WHERE d.dealership_code = %s AND s.captured_at = %s
"""

# Model-level statistics come from vw_vehicle_listing_model_mix, which owns
# KPI-LST-009..011 and is the only view publishing minimum and maximum advertised price.
# It is grained one level finer than this sheet -- it carries trim -- so the sheet rolls
# trim up here rather than showing one row per trim, which is what the sanitized workbook's
# own Model Summary does.
_MODEL_SQL = """
    SELECT m.condition_type,
           m.make,
           m.model,
           sum(m.observed_listing_units)::bigint          AS observed_listing_units,
           sum(m.listed_price_units)::bigint              AS listed_price_units,
           sum(m.call_for_price_units)::bigint            AS call_for_price_units,
           sum(m.unpriced_units)::bigint                  AS unpriced_units,
           sum(m.no_odometer_units)::bigint               AS no_odometer_units,
           CASE WHEN sum(m.listed_price_units) = 0 THEN NULL
                ELSE (sum(m.total_advertised_value) / sum(m.listed_price_units))::numeric(12,2)
           END                                            AS average_advertised_price,
           min(m.minimum_advertised_price)                AS minimum_advertised_price,
           max(m.maximum_advertised_price)                AS maximum_advertised_price,
           -- NULL when no listing in the rolled-up group published a reading. Left as
           -- NULL rather than coalesced to zero: a zero-mile row in a used-car report
           -- reads as a brand-new car, which is the opposite of what NULL means here.
           round(avg(m.average_odometer_miles))::bigint   AS average_odometer_miles
    FROM reporting.vw_vehicle_listing_model_mix AS m
    JOIN reporting.vw_dealership AS d ON d.dealership_key = m.dealership_key
    WHERE d.dealership_code = %s AND m.captured_at = %s
    GROUP BY m.condition_type, m.make, m.model
    ORDER BY (m.condition_type <> 'New'), sum(m.observed_listing_units) DESC,
             m.make, m.model
"""

_CHANGE_SQL = """
    SELECT ch.change_type, ch.condition_type, ch.make, ch.model, ch.trim,
           ch.vehicle_display, ch.synthetic_vehicle_id, ch.prior_advertised_price,
           ch.advertised_price, ch.price_change, ch.prior_pricing_status,
           ch.pricing_status, ch.prior_captured_at, ch.days_between_snapshots
    FROM reporting.vw_vehicle_listing_change AS ch
    JOIN reporting.vw_dealership AS d ON d.dealership_key = ch.dealership_key
    WHERE d.dealership_code = %s AND ch.captured_at = %s AND ch.has_prior_snapshot
    ORDER BY ch.change_type, ch.make, ch.model, ch.synthetic_vehicle_id
"""

_PRIOR_CAPTURE_SQL = """
    SELECT max(s.captured_at)
    FROM reporting.vw_vehicle_listing_summary AS s
    JOIN reporting.vw_dealership AS d ON d.dealership_key = s.dealership_key
    WHERE d.dealership_code = %s AND s.captured_at < %s
"""


def _rows(cursor: Any, statement: str, parameters: Sequence[Any]) -> list[tuple[Any, ...]]:
    """Run a query and return every row."""
    cursor.execute(statement, tuple(parameters))
    return list(cursor.fetchall())


# --------------------------------------------------------------------------------------
# Sheets
# --------------------------------------------------------------------------------------


def _write_readme(  # noqa: PLR0913 - the README block is one statement with many fields
    sheet: Any,
    *,
    dealership_id: str,
    store_name: str,
    market_region: str,
    captured_at: date,
    prior_captured_at: date | None,
    row_count: int,
    source_file_name: str,
    contract: InventoryListingContract,
) -> None:
    """Write the governance sheet: what this is, where it came from, what it cannot say."""
    write_title(sheet, "ARPI Inventory Operations | Listing Operating Report")
    facts = (
        ("Report", f"{store_name} observed listing operating report"),
        ("Dealer group", "Granite Auto Group"),
        ("Store", store_name),
        ("Dealership ID", dealership_id),
        ("Market", market_region),
        ("Capture date", captured_at),
        ("Prior capture date", prior_captured_at if prior_captured_at else "none"),
        ("Observed listings", row_count),
        ("Classification", contract.classification),
        ("Contract version", contract.contract_version),
        ("Generated from", "reporting.vw_vehicle_listing_* over PostgreSQL"),
        ("Source workbook", source_file_name),
    )
    for offset, (label, value) in enumerate(facts):
        sheet.cell(row=4 + offset, column=1, value=label)
        cell = sheet.cell(row=4 + offset, column=2, value=value)
        if isinstance(value, date):
            cell.number_format = DATE_FORMAT
    style_key_value_block(sheet, 4, 3 + len(facts))

    row = 5 + len(facts)
    write_heading(sheet, row, "What this report can and cannot establish")
    # Phrased as QUESTIONS, not as assertions. A cell reading "these vehicles were sold"
    # beside a cell reading "No" is unambiguous to a person and is exactly the sentence a
    # screenshot, a copy-paste or an automated content check would quote out of context.
    claims = (
        ("Question", "Answer", "Why"),
        (
            "Were these vehicles advertised on the capture date?",
            "Yes",
            "Each row is one listing observed in a de-identified public source on that date.",
        ),
        (
            "Were they physically on the ground?",
            "No",
            "A listing does not prove physical presence, and this source carries no lot record.",
        ),
        (
            "Did the dealership own them?",
            "No",
            "A listing does not prove ownership. Owned inventory is a different fact table.",
        ),
        (
            "Is the advertised price what a vehicle sold for?",
            "No",
            "Advertised price is not transaction price, acquisition cost, investment or gross.",
        ),
        (
            "Does a vehicle that stopped appearing mean it was sold?",
            "No",
            "A removed listing can reflect a sale, a trade, a wholesale, feed suppression or an "
            "error. This data cannot tell them apart, so nothing here says sold.",
        ),
        (
            "Can days observed online be read as days in stock?",
            "No",
            "Days in stock runs from acquisition and lives in the DMS. This lane never sees it.",
        ),
        (
            "Does this establish current business performance?",
            "No",
            "A public reference snapshot describes what was advertised, not how a store performed.",
        ),
    )
    for offset, claim in enumerate(claims):
        for column, value in enumerate(claim, start=1):
            sheet.cell(row=row + 1 + offset, column=column, value=value)
    style_table(
        sheet,
        name="ReportClaims",
        first_row=row + 1,
        last_row=row + len(claims),
        last_column=3,
        freeze=False,
        autofilter=False,
    )

    write_notice(
        sheet, row + len(claims) + 3, "Sanitized public reference notice", PUBLIC_REFERENCE_NOTICE
    )
    column_widths(sheet, (28, 56, 74))


def _write_inventory(
    sheet: Any, rows: Sequence[tuple[Any, ...]], contract: InventoryListingContract
) -> None:
    """Write the listing detail sheet, styled and filterable."""
    headers = (
        "Condition",
        "Model Year",
        "Make",
        "Model",
        "Trim",
        "Vehicle Display",
        "Odometer Miles",
        "Advertised Price",
        "Pricing Status",
        "Synthetic Vehicle ID",
        "Synthetic VIN",
        "Inventory Unit Count",
        "First Observed",
        "Last Observed",
        "Source Batch ID",
        "Source File Name",
    )
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column, value=header)
    for offset, row in enumerate(rows):
        for column, value in enumerate(row, start=1):
            cell = sheet.cell(row=2 + offset, column=column, value=value)
            if column in _INVENTORY_INTEGER_COLUMNS:
                cell.number_format = INTEGER_FORMAT
            elif column == _INVENTORY_CURRENCY_COLUMN:
                cell.number_format = CURRENCY_FORMAT
            elif column in _INVENTORY_DATE_COLUMNS:
                cell.number_format = DATE_FORMAT
    style_table(
        sheet,
        name="ListingDetail",
        first_row=1,
        last_row=1 + len(rows),
        last_column=len(headers),
    )
    column_widths(sheet, (11, 12, 18, 30, 18, 42, 15, 17, 15, 22, 20, 19, 15, 15, 22, 52))
    # `contract` is accepted so the signature matches its sibling writers and a future
    # contract change reaches this sheet without a signature churn.
    _ = contract


def _write_summary(  # noqa: PLR0913, PLR0915 - one sheet, written cell by cell
    # A sheet layout is a sequence of placements. Extracting each block into a helper
    # would hide the row arithmetic that makes the layout readable at all.
    sheet: Any,
    *,
    store_name: str,
    dealership_id: str,
    market_region: str,
    captured_at: date,
    row_count: int,
    summary_row: tuple[Any, ...] | None,
    top_models: Sequence[tuple[Any, ...]],
    contract: InventoryListingContract,
) -> None:
    """Write the formula-driven Summary sheet, with the condition and top-model charts."""
    last = row_count + 1
    detail = contract.sheets["inventory"]
    condition_range = f"'{detail}'!$A$2:$A${last}"
    make_range = f"'{detail}'!$C$2:$C${last}"
    model_range = f"'{detail}'!$D$2:$D${last}"
    price_range = f"'{detail}'!$H$2:$H${last}"
    status_range = f"'{detail}'!$I$2:$I${last}"
    id_range = f"'{detail}'!$J$2:$J${last}"

    write_title(sheet, f"ARPI Inventory Operations | {store_name}")
    facts = (
        ("Dealer group", "Granite Auto Group"),
        ("Store", store_name),
        ("Dealership ID", dealership_id),
        ("Market", market_region),
        ("Capture date", captured_at),
    )
    for offset, (label, value) in enumerate(facts):
        sheet.cell(row=4 + offset, column=1, value=label)
        cell = sheet.cell(row=4 + offset, column=2, value=value)
        if isinstance(value, date):
            cell.number_format = DATE_FORMAT
    style_key_value_block(sheet, 4, 3 + len(facts))

    sheet.cell(row=4, column=4, value="Condition")
    sheet.cell(row=4, column=5, value="Count")
    condition_rows = (
        ("New", f'=COUNTIF({condition_range},"New")'),
        ("Used", f'=COUNTIF({condition_range},"Used")'),
    )
    for offset, (label, formula) in enumerate(condition_rows):
        sheet.cell(row=5 + offset, column=4, value=label)
        sheet.cell(row=5 + offset, column=5, value=formula).number_format = INTEGER_FORMAT
    style_table(
        sheet,
        name="ReportConditionMix",
        first_row=4,
        last_row=4 + len(condition_rows),
        first_column=4,
        last_column=5,
        freeze=False,
        autofilter=False,
    )

    sheet.cell(row=10, column=1, value="Metric")
    sheet.cell(row=10, column=2, value="Value")

    # Metrics are declared as (label, formula, format) and the formulas that depend on
    # OTHER metrics reference them by label through `{units}`-style placeholders, filled
    # in below from each label's actual row.
    #
    # An earlier version wrote `B14` and `B16` literally. Inserting two rows above them
    # moved the values they pointed at, and the workbook went on opening cleanly while
    # dividing the wrong number by the right one -- a defect that survives every test
    # that does not evaluate the formulas. Positional references into a list somebody
    # will edit are a trap; this makes the reference say what it means.
    metric_specs: tuple[tuple[str, str, str], ...] = (
        ("Observed listing units", f"=COUNTA({id_range})", INTEGER_FORMAT),
        ("New listing units", f'=COUNTIF({condition_range},"New")', INTEGER_FORMAT),
        ("Used listing units", f'=COUNTIF({condition_range},"Used")', INTEGER_FORMAT),
        ("Vehicles with listed price", f'=COUNTIF({status_range},"Listed")', INTEGER_FORMAT),
        ("Call-for-price units", f'=COUNTIF({status_range},"Call for price")', INTEGER_FORMAT),
        (
            "Price-not-exposed units",
            f'=COUNTIF({status_range},"Price not exposed")',
            INTEGER_FORMAT,
        ),
        # The complement, so listed + unpriced is every row whatever statuses exist.
        (
            "Units with no listed price",
            f'=COUNTA({id_range})-COUNTIF({status_range},"Listed")',
            INTEGER_FORMAT,
        ),
        ("Total advertised listing value", f"=SUM({price_range})", CURRENCY_FORMAT),
        (
            "Average advertised price",
            "=IF({Vehicles with listed price}=0,0,"
            "{Total advertised listing value}/{Vehicles with listed price})",
            CURRENCY_FORMAT,
        ),
        (
            "Minimum advertised price",
            f"=IF({{Vehicles with listed price}}=0,0,MIN({price_range}))",
            CURRENCY_FORMAT,
        ),
        (
            "Maximum advertised price",
            f"=IF({{Vehicles with listed price}}=0,0,MAX({price_range}))",
            CURRENCY_FORMAT,
        ),
        (
            "Pricing completeness",
            "=IF({Observed listing units}=0,0,"
            "{Vehicles with listed price}/{Observed listing units})",
            PERCENT_FORMAT,
        ),
    )
    cell_of = {label: f"B{11 + index}" for index, (label, _, _) in enumerate(metric_specs)}

    def _resolve(formula: str) -> str:
        """Replace every ``{Metric label}`` with that metric's own cell reference."""
        for label, reference in cell_of.items():
            formula = formula.replace("{" + label + "}", reference)
        if "{" in formula:
            unresolved = formula[formula.index("{") :].split("}")[0] + "}"
            raise KeyError(f"Summary formula references an unknown metric: {unresolved}")
        return formula

    metrics = tuple(
        (label, _resolve(formula), number_format) for label, formula, number_format in metric_specs
    )
    for offset, (label, formula, number_format) in enumerate(metrics):
        sheet.cell(row=11 + offset, column=1, value=label)
        sheet.cell(row=11 + offset, column=2, value=formula).number_format = number_format
    style_table(
        sheet,
        name="ReportMetrics",
        first_row=10,
        last_row=10 + len(metrics),
        last_column=2,
        freeze=False,
        autofilter=False,
    )

    # The warehouse's own answer, beside the formulas that recompute it from the detail
    # sheet. If the two ever disagree, the workbook says so on its face rather than
    # requiring somebody to re-run the query.
    if summary_row is not None:
        sheet.cell(row=10, column=4, value="Warehouse figure")
        sheet.cell(row=10, column=5, value="Value")
        warehouse_metrics = (
            ("Observed listing units", summary_row[0], INTEGER_FORMAT),
            ("New listing units", summary_row[1], INTEGER_FORMAT),
            ("Used listing units", summary_row[2], INTEGER_FORMAT),
            ("Vehicles with listed price", summary_row[3], INTEGER_FORMAT),
            ("Call-for-price units", summary_row[4], INTEGER_FORMAT),
            ("Total advertised listing value", summary_row[5], CURRENCY_FORMAT),
            ("Snapshot age in days", summary_row[8], INTEGER_FORMAT),
        )
        for offset, (label, value, number_format) in enumerate(warehouse_metrics):
            sheet.cell(row=11 + offset, column=4, value=label)
            sheet.cell(row=11 + offset, column=5, value=value).number_format = number_format
        style_table(
            sheet,
            name="WarehouseFigures",
            first_row=10,
            last_row=10 + len(warehouse_metrics),
            first_column=4,
            last_column=5,
            freeze=False,
            autofilter=False,
        )

    row = 11 + len(metrics) + 1
    write_heading(sheet, row, "Governed interpretation")
    cell = sheet.cell(row=row + 1, column=1, value=GOVERNED_INTERPRETATION)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=6)
    sheet.row_dimensions[row + 1].height = 60

    row = row + 4
    write_heading(sheet, row, "Top observed models")
    sheet.cell(row=row + 1, column=1, value="Condition")
    sheet.cell(row=row + 1, column=2, value="Make")
    sheet.cell(row=row + 1, column=3, value="Model")
    sheet.cell(row=row + 1, column=4, value="Units")
    for offset, model_row in enumerate(top_models):
        target = row + 2 + offset
        sheet.cell(row=target, column=1, value=model_row[0])
        sheet.cell(row=target, column=2, value=model_row[1])
        sheet.cell(row=target, column=3, value=model_row[2])
        sheet.cell(
            row=target,
            column=4,
            value=(
                f"=COUNTIFS({condition_range},A{target},{make_range},B{target},"
                f"{model_range},C{target})"
            ),
        ).number_format = INTEGER_FORMAT
    top_last = row + 1 + len(top_models)
    style_table(
        sheet,
        name="ReportTopModels",
        first_row=row + 1,
        last_row=top_last,
        last_column=4,
        freeze=False,
        autofilter=False,
    )

    if len(top_models) >= 1:
        add_condition_chart(
            sheet,
            anchor="G4",
            labels="$D$5:$D$6",
            values="$E$4:$E$6",
            title="Observed listings by condition",
        )
        add_top_model_chart(
            sheet,
            anchor=f"G{row + 1}",
            labels=f"$C${row + 2}:$C${top_last}",
            values=f"$D${row + 1}:$D${top_last}",
            title="Top observed models by unit count",
        )

    write_notice(
        sheet,
        top_last + 3,
        "Sanitized public reference notice",
        PUBLIC_REFERENCE_NOTICE,
    )
    column_widths(sheet, (32, 26, 32, 30, 18, 18))


def _write_model_summary(
    sheet: Any, rows: Sequence[tuple[Any, ...]], contract: InventoryListingContract
) -> None:
    """Write the formula-driven Model Summary sheet."""
    detail = contract.sheets["inventory"]
    # The Inventory sheet holds one row per observed listing, so its last populated row is
    # the total unit count plus the header. Computed once: recomputing it per group would
    # produce the same number and invite the two to drift apart.
    detail_last = _detail_last_row(rows)
    condition_range = f"'{detail}'!$A$2:$A${detail_last}"
    make_range = f"'{detail}'!$C$2:$C${detail_last}"
    model_range = f"'{detail}'!$D$2:$D${detail_last}"
    price_range = f"'{detail}'!$H$2:$H${detail_last}"
    status_range = f"'{detail}'!$I$2:$I${detail_last}"

    write_title(sheet, "ARPI Inventory Operations | Model Summary")
    headers = (
        "Condition",
        "Make",
        "Model",
        "Total Units",
        "Listed Units",
        "No Listed Price",
        "Average Listed Price",
        "Minimum Listed Price",
        "Maximum Listed Price",
        "Average Odometer",
    )
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=4, column=column, value=header)

    for offset, model_row in enumerate(rows):
        target = 5 + offset
        for column, value in enumerate(model_row[:3], start=1):
            sheet.cell(row=target, column=column, value=value)
        criteria = f"{condition_range},$A{target},{make_range},$B{target},{model_range},$C{target}"
        sheet.cell(
            row=target, column=4, value=f"=COUNTIFS({criteria})"
        ).number_format = INTEGER_FORMAT
        sheet.cell(
            row=target, column=5, value=f'=COUNTIFS({criteria},{status_range},"Listed")'
        ).number_format = INTEGER_FORMAT
        # Total minus listed, so every unpriced status is counted whatever it is called.
        sheet.cell(
            row=target, column=6, value=f"=D{target}-E{target}"
        ).number_format = INTEGER_FORMAT
        sheet.cell(
            row=target,
            column=7,
            value=(
                f'=IF(E{target}=0,"",SUMIFS({price_range},{criteria},{status_range},'
                f'"Listed")/E{target})'
            ),
        ).number_format = CURRENCY_FORMAT
        sheet.cell(
            row=target,
            column=8,
            value=f'=IF(E{target}=0,"",MINIFS({price_range},{criteria},{status_range},"Listed"))',
        ).number_format = CURRENCY_FORMAT
        sheet.cell(
            row=target,
            column=9,
            value=f'=IF(E{target}=0,"",MAXIFS({price_range},{criteria},{status_range},"Listed"))',
        ).number_format = CURRENCY_FORMAT
        # Average odometer comes from the warehouse rather than a formula: the detail
        # sheet holds one row per vehicle, so a COUNTIFS-weighted mean over it would be
        # the same number computed less legibly, and the point of publishing it is to
        # show the warehouse's answer.
        sheet.cell(row=target, column=10, value=model_row[9]).number_format = INTEGER_FORMAT

    style_table(
        sheet,
        name="ReportModelSummary",
        first_row=4,
        last_row=4 + len(rows),
        last_column=len(headers),
    )
    column_widths(sheet, (12, 20, 34, 13, 13, 15, 20, 20, 20, 18))


def _detail_last_row(rows: Sequence[tuple[Any, ...]]) -> int:
    """Return the last populated row of the Inventory sheet the formulas point at."""
    return sum(int(row[3]) for row in rows) + 1


def _write_changes(
    sheet: Any, rows: Sequence[tuple[Any, ...]], *, prior_captured_at: date | None
) -> None:
    """Write the Snapshot Changes sheet, which exists only when a prior capture does."""
    write_title(sheet, "ARPI Inventory Operations | Snapshot Changes")
    sheet.cell(row=2, column=1, value="Compared against the prior capture on")
    cell = sheet.cell(row=2, column=2, value=prior_captured_at)
    cell.number_format = DATE_FORMAT
    style_key_value_block(sheet, 2, 2)

    headers = (
        "Change Type",
        "Condition",
        "Make",
        "Model",
        "Trim",
        "Vehicle Display",
        "Synthetic Vehicle ID",
        "Prior Advertised Price",
        "Advertised Price",
        "Price Change",
        "Prior Pricing Status",
        "Pricing Status",
        "Prior Capture Date",
        "Days Between Snapshots",
    )
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=4, column=column, value=header)
    for offset, row in enumerate(rows):
        for column, value in enumerate(row, start=1):
            written = sheet.cell(row=5 + offset, column=column, value=value)
            if column in _CHANGE_CURRENCY_COLUMNS:
                written.number_format = CURRENCY_FORMAT
            elif column == _CHANGE_DATE_COLUMN:
                written.number_format = DATE_FORMAT
            elif column == _CHANGE_INTEGER_COLUMN:
                written.number_format = INTEGER_FORMAT
    style_table(
        sheet,
        name="SnapshotChanges",
        first_row=4,
        last_row=4 + len(rows),
        last_column=len(headers),
    )
    column_widths(sheet, (22, 11, 18, 30, 18, 42, 22, 20, 18, 15, 20, 16, 18, 20))

    write_notice(
        sheet,
        6 + len(rows),
        "Removed From Listing is not Sold",
        "A listing can disappear because the vehicle sold, was traded to another store, was "
        "wholesaled, was suppressed by the feed, or because of an error. This data cannot "
        "distinguish those, so no row on this sheet says sold, delivered or disposed. "
        "Treat Removed From Listing as a prompt to look, not as a transaction.",
        width=14,
    )


# --------------------------------------------------------------------------------------
# Entry point
# --------------------------------------------------------------------------------------


def export_operating_report(
    connection: Any,
    *,
    dealership_id: str,
    captured_at: date,
    output_path: Path | None = None,
    contract: InventoryListingContract | None = None,
    overwrite: bool = True,
) -> ReportSummary:
    """Export the Excel operating report for one store and one capture date.

    Args:
        connection: Open ``psycopg`` connection with read access to ``reporting``.
        dealership_id: Store to report on.
        captured_at: Capture date to report as of.
        output_path: Where to write. Defaults to
            ``artifacts/inventory/ARPI_<Store_Descriptor>_Inventory_Report_<date>.xlsx``.
        contract: Contract supplying sheet names and the naming convention.
        overwrite: Whether an existing report may be replaced. A generated report is
            reproducible from the warehouse, so this defaults to ``True`` -- unlike the
            sanitizer, whose output is a committed artifact and defaults to refusal.

    Returns:
        A summary of what was produced.

    Raises:
        ValidationError: If the warehouse holds no listing rows for that store and capture,
            or if the output exists and ``overwrite`` is not set.
    """
    active = contract or load_contract()
    store = resolve_store(dealership_id)
    destination = output_path or (
        Path(active.report_root)
        / derived_report_file_name(store.dealership_id, captured_at, contract=active)
    )
    expected_name = derived_report_file_name(store.dealership_id, captured_at, contract=active)

    if destination.exists() and not overwrite:
        raise ValidationError(
            f"{destination.as_posix()} already exists and --overwrite was not given.",
            field="output",
        )

    with connection.cursor() as cursor:
        listings = _rows(cursor, _LISTING_SQL, (store.dealership_id, captured_at))
        if not listings:
            raise ValidationError(
                f"The warehouse holds no observed listings for {store.dealership_id} "
                f"captured on {captured_at.isoformat()}. Import the sanitized workbook "
                "first; an operating report over no rows would be indistinguishable from "
                "a store with no inventory.",
                field="captured_at",
            )
        summary_rows = _rows(cursor, _SUMMARY_SQL, (store.dealership_id, captured_at))
        model_rows = _rows(cursor, _MODEL_SQL, (store.dealership_id, captured_at))
        prior = _rows(cursor, _PRIOR_CAPTURE_SQL, (store.dealership_id, captured_at))
        prior_captured_at = prior[0][0] if prior and prior[0][0] else None
        change_rows = (
            _rows(cursor, _CHANGE_SQL, (store.dealership_id, captured_at))
            if prior_captured_at
            else []
        )

    sheets: list[str] = [
        active.sheets["readme"],
        active.sheets["summary"],
        active.sheets["inventory"],
        active.sheets["model_summary"],
    ]
    if prior_captured_at:
        sheets.append(active.optional_sheets["snapshot_changes"])

    with atomic_workbook_save(destination) as book:
        _write_readme(
            book.create_sheet(active.sheets["readme"]),
            dealership_id=store.dealership_id,
            store_name=store.store_name,
            market_region=store.market_region,
            captured_at=captured_at,
            prior_captured_at=prior_captured_at,
            row_count=len(listings),
            source_file_name=str(listings[0][15]),
            contract=active,
        )
        _write_summary(
            book.create_sheet(active.sheets["summary"]),
            store_name=store.store_name,
            dealership_id=store.dealership_id,
            market_region=store.market_region,
            captured_at=captured_at,
            row_count=len(listings),
            summary_row=summary_rows[0] if summary_rows else None,
            top_models=model_rows[:10],
            contract=active,
        )
        _write_inventory(book.create_sheet(active.sheets["inventory"]), listings, active)
        _write_model_summary(book.create_sheet(active.sheets["model_summary"]), model_rows, active)
        if prior_captured_at:
            _write_changes(
                book.create_sheet(active.optional_sheets["snapshot_changes"]),
                change_rows,
                prior_captured_at=prior_captured_at,
            )

    return ReportSummary(
        dealership_id=store.dealership_id,
        store_name=store.store_name,
        captured_at=captured_at,
        prior_captured_at=prior_captured_at,
        row_count=len(listings),
        model_group_count=len(model_rows),
        change_row_count=len(change_rows),
        sheets=tuple(sheets),
        output_path=destination,
        output_file_name=expected_name,
    )
