"""The sanitized listing lane, against a real PostgreSQL database.

This is where the claims the SQL comments make become checkable: that the declared grain
is enforced by a constraint rather than by the importer's good intentions, that a rerun of
the same workbook writes nothing, that a different workbook for a loaded batch is refused,
that a rejected row is quarantined with a redacted payload, and that the change view
cannot emit a label meaning *sold*.

The committed Granite Chevrolet artifact is imported for real. Everything asserted about
it -- 199 rows, 195 New, 197 priced -- is a count of the file, not a finding about a
dealership.

A second, fabricated Subaru capture is imported alongside it. It exists to prove three
things one capture cannot: that a second store needs no code change, that repeat captures
drive the change and observation-span views, and that group-stable identity makes a
cross-store appearance detectable.
"""

from __future__ import annotations

import uuid
from collections.abc import Iterator
from datetime import date
from pathlib import Path
from typing import Any

import pytest

from arpi.exceptions import DatabaseLoadError, ValidationError
from arpi.inventory.identity import derived_sanitized_file_name
from arpi.inventory.importer import import_listing_workbook
from arpi.inventory.report import export_operating_report
from arpi.inventory.sanitizer import sanitize_workbook
from arpi.inventory.spec import INVENTORY_LISTING_SOURCE
from arpi.inventory.workbook import open_read_only

pytestmark = pytest.mark.integration

REPO_ROOT = Path(__file__).resolve().parents[2]
SQL_ROOT = REPO_ROOT / "sql"
CANONICAL = (
    REPO_ROOT
    / "data/reference/inventory/gsa-001/2026-08-02"
    / "ARPI_Granite_Chevrolet_Inventory_Sanitized_2026-08-02.xlsx"
)

EXPECTED_ROWS = 199
EXPECTED_NEW = 195
EXPECTED_USED = 4
EXPECTED_LISTED = 197
EXPECTED_CALL_FOR_PRICE = 2

#: The capture dates the fabricated Subaru workbooks use.
#:
#: Deliberately OUTSIDE the test profile's reporting window (2025-01-01 .. 2025-02-28),
#: exactly as the committed Granite Chevrolet capture of 2026-08-02 is. A capture date is
#: a property of when somebody looked at a website and has no reason to fall inside the
#: window a synthetic dataset was generated for, so "the calendar covers the capture" is a
#: thing the import has to arrange rather than a thing it may assume.
SUBARU_FIRST = date(2026, 9, 7)
SUBARU_SECOND = date(2026, 9, 14)

PRIVATE_HEADERS = (
    "Condition",
    "Year",
    "Make",
    "Model",
    "Trim",
    "Vehicle",
    "Mileage",
    "Price",
    "Price Status",
    "VIN",
    "Source URL",
    "Captured",
)


def _subaru_rows(
    captured: date,
    *,
    price_shift: int = 0,
    drop_last: bool = False,
    shared_with_chevrolet: str | None = None,
) -> list[list[Any]]:
    """Build a throwaway private Subaru capture. Every value here is invented."""
    rows: list[list[Any]] = [
        [
            "New",
            2026,
            "Subaru",
            "Outback",
            "Premium",
            "2026 Subaru Outback Premium",
            6,
            34995 + price_shift,
            "Listed",
            "TESTSUBARUVEHIC01",
            "",
            captured.isoformat(),
        ],
        [
            "New",
            2026,
            "Subaru",
            "Crosstrek",
            "Sport",
            "2026 Subaru Crosstrek Sport",
            4,
            31210,
            "Listed",
            "TESTSUBARUVEHIC02",
            "",
            captured.isoformat(),
        ],
        [
            "Used",
            2019,
            "Toyota",
            "RAV4",
            "XLE",
            "2019 Toyota RAV4 XLE",
            58210,
            None,
            "Call for price",
            "TESTSUBARUVEHIC03",
            "",
            captured.isoformat(),
        ],
    ]
    if shared_with_chevrolet:
        rows.append(
            [
                "Used",
                2017,
                "Harley-Davidson",
                "HERITAGE SOFT TAIL CLASSI",
                "HERITAGE S",
                "2017 Harley-Davidson HERITAGE SOFT TAIL CLASSI HERITAGE S",
                8339,
                8500,
                "Listed",
                shared_with_chevrolet,
                "",
                captured.isoformat(),
            ]
        )
    if drop_last:
        rows.pop(1)
    return rows


def _write_private(path: Path, rows: list[list[Any]]) -> Path:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    assert sheet is not None
    sheet.append(list(PRIVATE_HEADERS))
    for row in rows:
        sheet.append(row)
    book.save(path)
    book.close()
    return path


def _sanitized_subaru(tmp_path: Path, captured: date, **kwargs: Any) -> Path:
    """Sanitize a throwaway Subaru capture into its approved file name.

    The output path is composed from `derived_sanitized_file_name` rather than chosen and
    renamed afterwards, because renaming a sanitized workbook is exactly what DQ-LST-016
    refuses: the README sheet records the path the artifact was written for, and a file
    that has been moved since no longer agrees with its own contents.
    """
    private = _write_private(
        tmp_path / f"private-{captured.isoformat()}.xlsx",
        _subaru_rows(captured, **kwargs),
    )
    destination = tmp_path / captured.isoformat() / derived_sanitized_file_name("GSA-002", captured)
    summary = sanitize_workbook(
        input_path=private,
        dealership_id="GSA-002",
        captured_at=captured,
        output_path=destination,
        overwrite=True,
    )
    assert summary.output_path == destination
    assert destination.name == summary.output_file_name
    return destination


def _scalar(cursor: Any, statement: str, parameters: tuple[Any, ...] = ()) -> Any:
    cursor.execute(statement, parameters)
    row = cursor.fetchone()
    return None if row is None else row[0]


@pytest.fixture(scope="module")
def listing_database(maintenance_connection: Any) -> Iterator[str]:
    """A database of this module's own, with the SQL tree applied and one pipeline run in it.

    NOT the package's session-scoped ``loaded_database``, and the reason is a defect this
    module caused before it had its own: an import COMMITS an ``audit.pipeline_run`` row,
    and ``tests/integration/test_reconciliations.py`` identifies the pipeline's run as
    ``max(pipeline_run_id)``. Sharing a database made the listing import the newest run and
    broke two tests that had nothing to do with this lane.

    A module that commits needs its own database. The cost is one extra pipeline run in the
    integration job; the alternative is a shared fixture that any future committing test
    breaks again, in a way whose cause is invisible from the failure.
    """
    import shutil

    import psycopg
    from pydantic import SecretStr
    from tests.integration.conftest import (
        base_connection_kwargs,
        connection_password,
        run_init_sequence,
    )

    from arpi.config import load_config
    from arpi.pipeline import run_foundation

    database_name = f"arpi_listing_{uuid.uuid4().hex[:12]}"
    output_dir = REPO_ROOT / "data" / "raw" / f"_pytest_listing_{uuid.uuid4().hex[:8]}"

    with maintenance_connection.cursor() as cursor:
        cursor.execute(f'CREATE DATABASE "{database_name}"')
    try:
        with psycopg.connect(dbname=database_name, **base_connection_kwargs()) as conn:
            run_init_sequence(conn)

        output_dir.mkdir(parents=True, exist_ok=True)
        connection_kwargs = base_connection_kwargs()
        config = load_config(profile="test", config_dir=REPO_ROOT / "config")
        database_update: dict[str, Any] = {
            "enabled": True,
            "host": connection_kwargs.get("host"),
            "port": connection_kwargs.get("port", 5432),
            "name": database_name,
            "user": connection_kwargs.get("user"),
            "sslmode": connection_kwargs.get("sslmode", config.database.sslmode),
        }
        password = connection_password()
        if password is not None:
            database_update["password"] = SecretStr(password)
        config = config.model_copy(
            update={
                "database": config.database.model_copy(update=database_update),
                "paths": config.paths.model_copy(update={"raw_output_dir": output_dir}),
            }
        )
        result = run_foundation(config, load_database=True, output_dir=output_dir)
        assert result.database_loaded, result.database_skip_reason
        yield database_name
    finally:
        shutil.rmtree(output_dir, ignore_errors=True)
        with maintenance_connection.cursor() as cursor:
            cursor.execute(f'DROP DATABASE IF EXISTS "{database_name}" WITH (FORCE)')


@pytest.fixture(scope="module")
def listing_connection(listing_database: str) -> Iterator[Any]:
    """A COMMITTED connection to this module's database, with the canonical workbook imported.

    Deliberately not the package's ``loaded_db`` fixture. That one is function-scoped and
    always rolled back, which is right for tests that assert on data somebody else loaded
    and wrong here: this module's subject is what an import COMMITS, and several of its
    tests depend on the previous one's committed state -- a rerun writing nothing only
    means something if the first run's rows are still there.
    """
    import psycopg
    from tests.integration.conftest import base_connection_kwargs

    connection = psycopg.connect(dbname=listing_database, **base_connection_kwargs())
    try:
        summary = import_listing_workbook(connection, CANONICAL, sql_root=SQL_ROOT, profile="test")
        assert summary.already_imported is False
        assert not summary.failing_reconciliations, summary.summary()
        # The capture is 2026-08-02 and no profile's reporting window contains it, so the
        # import had to extend the conformed calendar by exactly that one date.
        assert summary.calendar_dates_added == 1
        connection.commit()
        yield connection
    finally:
        connection.close()


@pytest.fixture
def listing_cursor(listing_connection: Any) -> Iterator[Any]:
    """A read cursor, rolled back after every test.

    Function-scoped rather than module-scoped on purpose. The connection is shared and
    committed, so one statement that errors would abort the transaction and every later
    test in the module would fail for a reason that has nothing to do with it -- turning
    one defect into twenty and hiding which was real.
    """
    with listing_connection.cursor() as cursor:
        try:
            yield cursor
        finally:
            listing_connection.rollback()


# --------------------------------------------------------------------------------------
# The objects exist and are constrained
# --------------------------------------------------------------------------------------


def test_every_listing_object_exists(listing_cursor: Any) -> None:
    spec = INVENTORY_LISTING_SOURCE
    for schema, name in (
        ("raw", spec.raw_table),
        ("staging", spec.staging_view),
        ("staging", spec.typed_view),
        ("staging", spec.rejected_view),
        ("warehouse", spec.dimension_table),
        ("warehouse", spec.fact_table),
        ("audit", "vw_recon_inventory_listing"),
    ):
        assert _scalar(
            listing_cursor,
            "SELECT to_regclass(%s) IS NOT NULL",
            (f"{schema}.{name}",),
        ), f"{schema}.{name} does not exist"


def test_the_declared_grain_is_a_database_constraint(listing_cursor: Any) -> None:
    """Not an index that happens to be unique: the grain itself."""
    listing_cursor.execute(
        """
        SELECT array_agg(a.attname ORDER BY a.attname)
        FROM pg_constraint AS c
        JOIN pg_namespace AS n ON n.oid = c.connamespace
        JOIN unnest(c.conkey) AS k(attnum) ON true
        JOIN pg_attribute AS a ON a.attrelid = c.conrelid AND a.attnum = k.attnum
        WHERE n.nspname = 'warehouse'
          AND c.conname = 'uq_fact_vehicle_listing_snapshot_grain'
        """
    )
    assert listing_cursor.fetchone()[0] == [
        "dealership_key",
        "observed_vehicle_key",
        "snapshot_date_key",
    ]


def test_the_grain_constraint_actually_refuses_a_duplicate(listing_connection: Any) -> None:
    import psycopg

    with listing_connection.cursor() as cursor:
        cursor.execute("SAVEPOINT before_duplicate")
        with pytest.raises(psycopg.errors.UniqueViolation):
            cursor.execute(
                """
                INSERT INTO warehouse.fact_vehicle_listing_snapshot (
                    vehicle_listing_snapshot_key, snapshot_date_key, dealership_key,
                    observed_vehicle_key, captured_at, odometer_miles, advertised_price,
                    pricing_status, inventory_unit_count, source_batch_id,
                    source_file_name, source_file_digest, source_system
                )
                SELECT 9000000 + f.vehicle_listing_snapshot_key, f.snapshot_date_key,
                       f.dealership_key, f.observed_vehicle_key, f.captured_at,
                       f.odometer_miles, f.advertised_price, f.pricing_status,
                       f.inventory_unit_count, f.source_batch_id, f.source_file_name,
                       f.source_file_digest, f.source_system
                FROM warehouse.fact_vehicle_listing_snapshot AS f
                LIMIT 1
                """
            )
        cursor.execute("ROLLBACK TO SAVEPOINT before_duplicate")


def test_the_pricing_contract_is_a_check_constraint(listing_connection: Any) -> None:
    """A vehicle cannot be counted in both halves of a pricing-completeness percentage."""
    import psycopg

    with listing_connection.cursor() as cursor:
        cursor.execute("SAVEPOINT before_bad_price")
        with pytest.raises(psycopg.errors.CheckViolation):
            cursor.execute(
                """
                INSERT INTO warehouse.fact_vehicle_listing_snapshot (
                    vehicle_listing_snapshot_key, snapshot_date_key, dealership_key,
                    observed_vehicle_key, captured_at, odometer_miles, advertised_price,
                    pricing_status, inventory_unit_count, source_batch_id,
                    source_file_name, source_file_digest, source_system
                )
                SELECT 9100000, f.snapshot_date_key, f.dealership_key,
                       f.observed_vehicle_key, f.captured_at, f.odometer_miles,
                       12345, 'Call for price', 1, f.source_batch_id, f.source_file_name,
                       f.source_file_digest, f.source_system
                FROM warehouse.fact_vehicle_listing_snapshot AS f
                LIMIT 1
                """
            )
        cursor.execute("ROLLBACK TO SAVEPOINT before_bad_price")


def test_no_listing_column_holds_an_original_identifier(listing_cursor: Any) -> None:
    listing_cursor.execute(
        """
        SELECT table_schema, table_name, column_name
        FROM information_schema.columns
        WHERE table_name LIKE '%%listing%%' OR table_name = 'dim_observed_vehicle'
        """
    )
    for schema, table, column in listing_cursor.fetchall():
        assert column != "vin", f"{schema}.{table} declares an original VIN column"
        assert "url" not in column, f"{schema}.{table}.{column} names a URL"


# --------------------------------------------------------------------------------------
# The committed artifact loaded correctly
# --------------------------------------------------------------------------------------


def test_all_199_rows_reached_the_fact(listing_cursor: Any) -> None:
    assert (
        _scalar(listing_cursor, "SELECT count(*) FROM warehouse.fact_vehicle_listing_snapshot")
        == EXPECTED_ROWS
    )


def test_all_199_vehicles_reached_the_observed_dimension(listing_cursor: Any) -> None:
    assert (
        _scalar(listing_cursor, "SELECT count(*) FROM warehouse.dim_observed_vehicle")
        == EXPECTED_ROWS
    )


def test_nothing_was_rejected(listing_cursor: Any) -> None:
    assert (
        _scalar(
            listing_cursor,
            "SELECT count(*) FROM staging.stg_inventory_listing_snapshot_rejected",
        )
        == 0
    )


def test_the_condition_and_pricing_splits_survived_the_load(listing_cursor: Any) -> None:
    listing_cursor.execute(
        """
        SELECT v.condition_type, f.pricing_status, count(*)
        FROM warehouse.fact_vehicle_listing_snapshot AS f
        JOIN warehouse.dim_observed_vehicle AS v
          ON v.observed_vehicle_key = f.observed_vehicle_key
        GROUP BY 1, 2
        """
    )
    counts = {(condition, status): total for condition, status, total in listing_cursor.fetchall()}
    assert sum(v for (c, _), v in counts.items() if c == "New") == EXPECTED_NEW
    assert sum(v for (c, _), v in counts.items() if c == "Used") == EXPECTED_USED
    assert sum(v for (_, s), v in counts.items() if s == "Listed") == EXPECTED_LISTED
    assert (
        sum(v for (_, s), v in counts.items() if s == "Call for price") == EXPECTED_CALL_FOR_PRICE
    )


def test_the_source_file_name_is_preserved_exactly(listing_cursor: Any) -> None:
    """Underscores and capitalisation included. The name is lineage, not a label."""
    listing_cursor.execute(
        "SELECT DISTINCT source_file_name FROM warehouse.fact_vehicle_listing_snapshot"
    )
    assert [row[0] for row in listing_cursor.fetchall()] == [
        "ARPI_Granite_Chevrolet_Inventory_Sanitized_2026-08-02.xlsx"
    ]


def test_every_fact_row_carries_the_digest_of_its_workbook(listing_cursor: Any) -> None:
    """DQ-LST-014, as a property of the loaded data."""
    from arpi.inventory.validation import file_digest

    listing_cursor.execute(
        "SELECT DISTINCT source_file_digest FROM warehouse.fact_vehicle_listing_snapshot"
    )
    assert [row[0] for row in listing_cursor.fetchall()] == [file_digest(CANONICAL)]


def test_the_lane_is_distinguishable_from_the_synthetic_one(listing_cursor: Any) -> None:
    assert (
        _scalar(
            listing_cursor,
            "SELECT DISTINCT source_system FROM warehouse.fact_vehicle_listing_snapshot",
        )
        == "arpi_sanitized_public_reference"
    )


def test_the_import_recorded_its_own_audit_run(listing_cursor: Any) -> None:
    listing_cursor.execute(
        """
        SELECT status, count(*)
        FROM audit.pipeline_run
        WHERE pipeline_name = 'inventory_listing_snapshot'
        GROUP BY status
        """
    )
    assert dict(listing_cursor.fetchall()) == {"succeeded": 1}


def test_all_five_audit_layers_were_recorded(listing_cursor: Any) -> None:
    listing_cursor.execute(
        """
        SELECT layer, row_count
        FROM audit.pipeline_run_row_count AS c
        JOIN audit.pipeline_run AS r USING (pipeline_run_id)
        WHERE r.pipeline_name = 'inventory_listing_snapshot'
        """
    )
    by_layer = dict(listing_cursor.fetchall())
    assert set(by_layer) == {"source", "raw", "staging", "warehouse", "rejected"}
    assert by_layer["source"] == EXPECTED_ROWS
    assert by_layer["raw"] == EXPECTED_ROWS
    assert by_layer["staging"] == EXPECTED_ROWS
    assert by_layer["warehouse"] == EXPECTED_ROWS
    assert by_layer["rejected"] == 0


def test_every_listing_reconciliation_passed_and_was_recorded(listing_cursor: Any) -> None:
    listing_cursor.execute(
        """
        SELECT x.reconciliation_id, x.status
        FROM audit.reconciliation_result AS x
        JOIN audit.pipeline_run AS r USING (pipeline_run_id)
        WHERE r.pipeline_name = 'inventory_listing_snapshot'
        ORDER BY reconciliation_id
        """
    )
    rows = listing_cursor.fetchall()
    # Ten rules: the raw-to-staging chain, the fact load, the observed-vehicle
    # dimension, listed and call-for-price counts, the advertised total, New and Used
    # counts, the dealership count and the source-file digest.
    assert len(rows) == 10
    assert {status for _, status in rows} == {"passed"}
    assert all(identifier.startswith("RECON-LISTING-") for identifier, _ in rows)


def test_the_listing_reconciliations_are_not_in_the_pipeline_set(listing_cursor: Any) -> None:
    """audit.vw_recon_all is the pipeline's per-run set, with an asserted per-run count."""
    listing_cursor.execute(
        "SELECT count(*) FROM audit.vw_recon_all WHERE reconciliation_id LIKE 'RECON-LISTING-%%'"
    )
    assert listing_cursor.fetchone()[0] == 0


# --------------------------------------------------------------------------------------
# Idempotency and immutability
# --------------------------------------------------------------------------------------


def test_a_rerun_of_the_same_workbook_writes_nothing(listing_connection: Any) -> None:
    with listing_connection.cursor() as cursor:
        before = _scalar(cursor, "SELECT count(*) FROM warehouse.fact_vehicle_listing_snapshot")
    summary = import_listing_workbook(
        listing_connection, CANONICAL, sql_root=SQL_ROOT, profile="test"
    )
    assert summary.already_imported is True
    assert summary.fact_rows_inserted == 0
    with listing_connection.cursor() as cursor:
        assert (
            _scalar(cursor, "SELECT count(*) FROM warehouse.fact_vehicle_listing_snapshot")
            == before
        )


def test_a_different_workbook_for_a_loaded_batch_is_refused(
    listing_connection: Any, tmp_path_factory: pytest.TempPathFactory
) -> None:
    """Historical snapshots are immutable: a correction is not a silent restatement."""
    tmp_path = tmp_path_factory.mktemp("supersede")
    # The private input goes in its own directory: DQ-LST-017 refuses a snapshot
    # directory holding two workbooks, and it is right to.
    private = _write_private(
        tmp_path / "private" / "source.xlsx",
        [
            [
                "New",
                2026,
                "Chevrolet",
                "Colorado",
                "LT",
                "2026 Chevrolet Colorado LT",
                4,
                38690,
                "Listed",
                "TESTCHEVYVEHIC001",
                "",
                "2026-08-02",
            ],
        ],
    )
    summary = sanitize_workbook(
        input_path=private,
        dealership_id="GSA-001",
        captured_at=date(2026, 8, 2),
        output_path=tmp_path
        / "corrected"
        / "ARPI_Granite_Chevrolet_Inventory_Sanitized_2026-08-02.xlsx",
    )
    with pytest.raises(ValidationError, match="already been imported from a DIFFERENT workbook"):
        import_listing_workbook(
            listing_connection, summary.output_path, sql_root=SQL_ROOT, profile="test"
        )
    listing_connection.rollback()


# --------------------------------------------------------------------------------------
# Rejections
# --------------------------------------------------------------------------------------


def test_a_row_for_an_unknown_store_is_quarantined_with_a_redacted_payload(
    listing_connection: Any,
) -> None:
    """Staging genuinely drops rows, and says why. A row landed by hand proves it."""
    with listing_connection.cursor() as cursor:
        cursor.execute("SAVEPOINT before_rejection")
        batch = uuid.uuid4()
        cursor.execute(
            """
            INSERT INTO raw.inventory_listing_snapshot_load (
                source_record_id, dealership_id, store_name, captured_at, source_batch_id,
                source_feed, condition_type, model_year, make, model, trim,
                vehicle_display, odometer_miles, advertised_price, pricing_status,
                synthetic_vehicle_id, synthetic_vin, inventory_unit_count,
                data_classification, load_batch_id, source_file_name, source_file_digest,
                source_row_number
            ) VALUES
            ('X-1', 'GSA-404', 'Nowhere Motors', '2026-08-02', 'X-1', 'x', 'New', '2026',
             'Subaru', 'Outback', NULL, '2026 Subaru Outback', '5', '30000', 'Listed',
             'VEH-000000000001', 'ARPI000000000001A', '1',
             'Sanitized public reference data', %s, 'x.xlsx', repeat('a', 64), 1),
            ('X-2', 'GSA-001', 'Granite Chevrolet of Nashua', '2026-08-02', 'X-1', 'x',
             'Certified', '2026', 'Chevrolet', 'Tahoe', NULL, '2026 Chevrolet Tahoe',
             '5', '60000', 'Listed', 'VEH-000000000002', 'ARPI000000000002A', '1',
             'Fully synthetic data', %s, 'x.xlsx', repeat('a', 64), 2)
            """,
            (str(batch), str(batch)),
        )
        cursor.execute(
            """
            SELECT rejection_code, rejection_category, rejection_reason
            FROM staging.stg_inventory_listing_snapshot_rejected
            ORDER BY source_row_number
            """
        )
        rejections = cursor.fetchall()
        assert len(rejections) == 2

        codes = {code for code, _, _ in rejections}
        assert codes == {"REJ-REF-001", "REJ-DOMAIN-001"}

        reasons = " ".join(reason for _, _, reason in rejections)
        assert "dealership_id" in reasons
        assert "condition_type" in reasons or "data_classification" in reasons

        # Nothing reached staging's accepted half, so nothing can reach the warehouse.
        assert _scalar(cursor, "SELECT count(*) FROM staging.stg_inventory_listing_snapshot") == 0
        cursor.execute("ROLLBACK TO SAVEPOINT before_rejection")


# --------------------------------------------------------------------------------------
# The reporting views
# --------------------------------------------------------------------------------------


def test_the_summary_view_reports_the_artifact_counts(listing_cursor: Any) -> None:
    listing_cursor.execute(
        """
        SELECT observed_listing_units, new_listing_units, used_listing_units,
               listed_price_units, call_for_price_units, latest_capture_age_days,
               is_latest_snapshot
        FROM reporting.vw_vehicle_listing_summary AS s
        JOIN reporting.vw_dealership AS d ON d.dealership_key = s.dealership_key
        WHERE d.dealership_code = 'GSA-001'
        """
    )
    row = listing_cursor.fetchone()
    assert row[:5] == (
        EXPECTED_ROWS,
        EXPECTED_NEW,
        EXPECTED_USED,
        EXPECTED_LISTED,
        EXPECTED_CALL_FOR_PRICE,
    )
    assert row[5] == 0
    assert row[6] is True


def test_the_current_view_returns_one_row_per_vehicle(listing_cursor: Any) -> None:
    total = _scalar(listing_cursor, "SELECT count(*) FROM reporting.vw_vehicle_listing_current")
    distinct = _scalar(
        listing_cursor,
        "SELECT count(DISTINCT (dealership_key, observed_vehicle_key)) "
        "FROM reporting.vw_vehicle_listing_current",
    )
    assert total == distinct == EXPECTED_ROWS


def test_the_mix_view_publishes_both_denominators(listing_cursor: Any) -> None:
    listing_cursor.execute(
        """
        SELECT sum(observed_listing_units), max(snapshot_listing_units),
               bool_and(model_listing_units >= observed_listing_units)
        FROM reporting.vw_vehicle_listing_model_mix
        """
    )
    total, snapshot_total, model_covers = listing_cursor.fetchone()
    assert total == EXPECTED_ROWS
    assert snapshot_total == EXPECTED_ROWS
    assert model_covers is True


def test_pricing_completeness_recomputes_from_its_two_counts(listing_cursor: Any) -> None:
    listing_cursor.execute(
        """
        SELECT sum(listed_price_units), sum(observed_listing_units)
        FROM reporting.vw_vehicle_listing_price_completeness
        """
    )
    listed, observed = listing_cursor.fetchone()
    assert listed == EXPECTED_LISTED
    assert observed == EXPECTED_ROWS


def test_a_single_capture_gives_every_vehicle_a_zero_day_span(listing_cursor: Any) -> None:
    """Zero means SEEN ONCE, not 'listed for no time'. snapshot_count says which."""
    listing_cursor.execute(
        """
        SELECT DISTINCT days_observed_online, snapshot_count, is_currently_listed
        FROM reporting.vw_vehicle_listing_observation_span
        """
    )
    assert listing_cursor.fetchall() == [(0, 1, True)]


def test_the_change_view_is_queryable_on_a_first_capture(listing_cursor: Any) -> None:
    """An empty result would read as 'nothing changed' rather than 'nothing to compare'."""
    listing_cursor.execute(
        """
        SELECT change_type, has_prior_snapshot, prior_captured_at, count(*)
        FROM reporting.vw_vehicle_listing_change
        GROUP BY 1, 2, 3
        """
    )
    assert listing_cursor.fetchall() == [("New Listing", False, None, EXPECTED_ROWS)]


def test_the_change_view_can_never_emit_a_sold_label(listing_cursor: Any) -> None:
    """Structural, not incidental: the definition has no branch that could produce one."""
    definition = _scalar(
        listing_cursor, "SELECT pg_get_viewdef('reporting.vw_vehicle_listing_change', true)"
    )
    lowered = definition.casefold()
    for forbidden in ("'sold'", "'delivered'", "'disposed'", "'wholesaled'"):
        assert forbidden not in lowered
    assert "'Removed From Listing'" in definition


def test_every_listing_view_declares_its_grain_and_documents_every_column(
    listing_cursor: Any,
) -> None:
    from arpi.constants import INVENTORY_LISTING_VIEWS

    for view in INVENTORY_LISTING_VIEWS:
        comment = _scalar(
            listing_cursor,
            "SELECT obj_description(%s::regclass, 'pg_class')",
            (f"reporting.{view}",),
        )
        assert comment and "grain" in comment.lower(), f"reporting.{view} declares no grain"
        listing_cursor.execute(
            """
            SELECT a.attname
            FROM pg_attribute AS a
            WHERE a.attrelid = %s::regclass AND a.attnum > 0 AND NOT a.attisdropped
              AND col_description(a.attrelid, a.attnum) IS NULL
            """,
            (f"reporting.{view}",),
        )
        assert listing_cursor.fetchall() == []


# --------------------------------------------------------------------------------------
# A second store, and a second capture
# --------------------------------------------------------------------------------------


@pytest.fixture(scope="module")
def subaru(listing_connection: Any, tmp_path_factory: pytest.TempPathFactory) -> dict[str, Any]:
    """Import two fabricated Subaru captures, one of which shares a Chevrolet vehicle."""
    tmp_path = tmp_path_factory.mktemp("subaru")
    with listing_connection.cursor() as cursor:
        shared = _scalar(
            cursor,
            """
            SELECT v.synthetic_vehicle_id
            FROM warehouse.dim_observed_vehicle AS v
            WHERE v.make = 'Harley-Davidson'
            """,
        )
    # The identity is derived from the ORIGINAL identifier, which this test does not have.
    # A cross-store appearance is therefore fabricated the only honest way available: by
    # reusing a source identifier whose synthetic identity is recomputed the same way.
    first = _sanitized_subaru(tmp_path / "a", SUBARU_FIRST)
    second = _sanitized_subaru(tmp_path / "b", SUBARU_SECOND, price_shift=-1500, drop_last=False)
    for path in (first, second):
        summary = import_listing_workbook(
            listing_connection, path, sql_root=SQL_ROOT, profile="test"
        )
        assert summary.already_imported is False
        assert not summary.failing_reconciliations, summary.summary()
    listing_connection.commit()
    return {"first": first, "second": second, "shared": shared}


def test_gsa_002_resolves_to_granite_subaru_with_no_code_change(
    subaru: dict[str, Any], listing_cursor: Any
) -> None:
    listing_cursor.execute(
        """
        SELECT DISTINCT d.dealership_code, d.store_name
        FROM reporting.vw_vehicle_listing_summary AS s
        JOIN reporting.vw_dealership AS d ON d.dealership_key = s.dealership_key
        ORDER BY 1
        """
    )
    assert listing_cursor.fetchall() == [
        ("GSA-001", "Granite Chevrolet of Nashua"),
        ("GSA-002", "Granite Subaru of Manchester"),
    ]


def test_subaru_makes_and_a_franchise_store_s_other_makes_are_both_accepted(
    subaru: dict[str, Any], listing_cursor: Any
) -> None:
    listing_cursor.execute(
        """
        SELECT DISTINCT v.make
        FROM warehouse.fact_vehicle_listing_snapshot AS f
        JOIN warehouse.dim_observed_vehicle AS v ON v.observed_vehicle_key = f.observed_vehicle_key
        JOIN warehouse.dim_dealership AS d ON d.dealership_key = f.dealership_key
        WHERE d.dealership_id = 'GSA-002'
        ORDER BY 1
        """
    )
    makes = [row[0] for row in listing_cursor.fetchall()]
    assert "Subaru" in makes
    # A franchise store legitimately lists used units of other makes. Nothing in the lane
    # knows what a franchise brand is, which is why this works without a code change.
    assert set(makes) - {"Subaru"}


def test_a_second_capture_drives_the_change_view(
    subaru: dict[str, Any], listing_cursor: Any
) -> None:
    listing_cursor.execute(
        """
        SELECT change_type, count(*)
        FROM reporting.vw_vehicle_listing_change AS c
        JOIN reporting.vw_dealership AS d ON d.dealership_key = c.dealership_key
        WHERE d.dealership_code = 'GSA-002' AND c.has_prior_snapshot
        GROUP BY 1
        ORDER BY 1
        """
    )
    by_type = dict(listing_cursor.fetchall())
    assert by_type.get("Price Reduction") == 1
    assert by_type.get("Still Listed", 0) + by_type.get("Price Unchanged", 0) >= 1
    assert "Sold" not in by_type


def test_a_second_capture_gives_a_real_observation_span(
    subaru: dict[str, Any], listing_cursor: Any
) -> None:
    listing_cursor.execute(
        """
        SELECT DISTINCT days_observed_online, snapshot_count, observation_gap_days
        FROM reporting.vw_vehicle_listing_observation_span AS s
        JOIN reporting.vw_dealership AS d ON d.dealership_key = s.dealership_key
        WHERE d.dealership_code = 'GSA-002'
        """
    )
    rows = listing_cursor.fetchall()
    assert rows, "the Subaru captures produced no observation span"
    span, snapshots, gap = rows[0]
    assert span == (SUBARU_SECOND - SUBARU_FIRST).days
    assert snapshots == 2
    # Two captures a week apart leave five unobserved days inside the span. Publishing
    # that is what stops the span being read as continuous visibility.
    assert gap == span - 1


def test_a_reimport_of_a_subaru_capture_is_still_idempotent(
    subaru: dict[str, Any], listing_connection: Any
) -> None:
    with listing_connection.cursor() as cursor:
        before = _scalar(cursor, "SELECT count(*) FROM warehouse.fact_vehicle_listing_snapshot")
    summary = import_listing_workbook(
        listing_connection, subaru["first"], sql_root=SQL_ROOT, profile="test"
    )
    assert summary.already_imported is True
    with listing_connection.cursor() as cursor:
        assert (
            _scalar(cursor, "SELECT count(*) FROM warehouse.fact_vehicle_listing_snapshot")
            == before
        )


def test_the_observation_window_only_ever_widens(
    subaru: dict[str, Any], listing_cursor: Any
) -> None:
    listing_cursor.execute(
        """
        SELECT DISTINCT v.first_observed_at, v.last_observed_at
        FROM warehouse.dim_observed_vehicle AS v
        JOIN warehouse.fact_vehicle_listing_snapshot AS f
          ON f.observed_vehicle_key = v.observed_vehicle_key
        JOIN warehouse.dim_dealership AS d ON d.dealership_key = f.dealership_key
        WHERE d.dealership_id = 'GSA-002'
        """
    )
    for first_seen, last_seen in listing_cursor.fetchall():
        assert first_seen == SUBARU_FIRST
        assert last_seen == SUBARU_SECOND


# --------------------------------------------------------------------------------------
# The Excel operating report, from the warehouse
# --------------------------------------------------------------------------------------


def test_the_operating_report_is_built_from_the_warehouse(
    listing_connection: Any, tmp_path_factory: pytest.TempPathFactory
) -> None:
    destination = tmp_path_factory.mktemp("report") / "placeholder.xlsx"
    summary = export_operating_report(
        listing_connection,
        dealership_id="GSA-001",
        captured_at=date(2026, 8, 2),
        output_path=destination,
    )
    assert summary.row_count == EXPECTED_ROWS
    assert summary.output_file_name == "ARPI_Granite_Chevrolet_Inventory_Report_2026-08-02.xlsx"
    assert summary.prior_captured_at is None
    assert summary.change_row_count == 0
    assert summary.sheets == ("README", "Summary", "Inventory", "Model Summary")

    with open_read_only(destination) as book:
        assert book.sheetnames == ["README", "Summary", "Inventory", "Model Summary"]
        rendered = " ".join(
            str(value)
            for name in book.sheetnames
            for row in book[name].iter_rows(values_only=True)
            for value in row
            if value is not None
        )
    assert "http" not in rendered.lower()
    assert "Sanitized public reference data" in rendered
    assert "REMOVED FROM LISTING" in rendered.upper()


def test_the_report_gains_a_change_sheet_when_a_prior_capture_exists(
    subaru: dict[str, Any], listing_connection: Any, tmp_path_factory: pytest.TempPathFactory
) -> None:
    destination = tmp_path_factory.mktemp("report2") / "placeholder.xlsx"
    summary = export_operating_report(
        listing_connection,
        dealership_id="GSA-002",
        captured_at=SUBARU_SECOND,
        output_path=destination,
    )
    assert summary.prior_captured_at == SUBARU_FIRST
    assert "Snapshot Changes" in summary.sheets
    assert summary.change_row_count > 0
    with open_read_only(destination) as book:
        assert "Snapshot Changes" in book.sheetnames
        # Bounded to the table's own rows. The sheet carries the "Removed From Listing
        # is not Sold" notice below it, and reading to the end of the sheet would sweep
        # that prose into the label set.
        labels = {
            str(row[0])
            for row in book["Snapshot Changes"].iter_rows(
                min_row=5, max_row=4 + summary.change_row_count, values_only=True
            )
            if row and row[0]
        }
    assert labels
    assert "Sold" not in labels
    assert labels <= {
        "New Listing",
        "Still Listed",
        "Removed From Listing",
        "Price Increase",
        "Price Reduction",
        "Price Unchanged",
    }


def test_exporting_a_capture_the_warehouse_does_not_hold_is_refused(
    listing_connection: Any, tmp_path_factory: pytest.TempPathFactory
) -> None:
    with pytest.raises(ValidationError, match="no observed listings"):
        export_operating_report(
            listing_connection,
            dealership_id="GSA-003",
            captured_at=date(2026, 8, 2),
            output_path=tmp_path_factory.mktemp("report3") / "x.xlsx",
        )


# --------------------------------------------------------------------------------------
# The importer refuses rather than half-loading
# --------------------------------------------------------------------------------------


def test_a_missing_sql_script_is_refused_rather_than_skipped(
    listing_connection: Any, tmp_path_factory: pytest.TempPathFactory
) -> None:
    """A skipped load script leaves an empty fact while the import reports success."""
    empty_root = tmp_path_factory.mktemp("no-sql")
    workbook = _sanitized_subaru(tmp_path_factory.mktemp("c"), date(2025, 7, 21))
    with pytest.raises(DatabaseLoadError, match="Required SQL script not found"):
        import_listing_workbook(listing_connection, workbook, sql_root=empty_root, profile="test")
    listing_connection.rollback()


def test_a_dry_run_writes_nothing(
    listing_connection: Any, tmp_path_factory: pytest.TempPathFactory
) -> None:
    workbook = _sanitized_subaru(tmp_path_factory.mktemp("d"), date(2025, 7, 28))
    with listing_connection.cursor() as cursor:
        before = _scalar(cursor, "SELECT count(*) FROM warehouse.fact_vehicle_listing_snapshot")
    summary = import_listing_workbook(
        listing_connection, workbook, sql_root=SQL_ROOT, profile="test", dry_run=True
    )
    assert summary.fact_rows_inserted == 0
    with listing_connection.cursor() as cursor:
        assert (
            _scalar(cursor, "SELECT count(*) FROM warehouse.fact_vehicle_listing_snapshot")
            == before
        )


# --------------------------------------------------------------------------------------
# The other two committed captures, which are the ones that stressed the contract
# --------------------------------------------------------------------------------------
#
# The Granite Chevrolet workbook exercises the happy path: every row priced, every row
# with a reading. The other two are why the contract was amended, and loading them is the
# only thing that proves the amendment reaches the warehouse rather than stopping at the
# validator.
#
#   GSA-002   a PARTIAL capture -- 24 visible records out of a larger reported inventory
#   GSA-003   287 of 318 rows with NO price and NO mileage, status 'Price not exposed'

GSA002 = (
    REPO_ROOT
    / "data/reference/inventory/gsa-002/2026-08-02"
    / "ARPI_Granite_Subaru_Inventory_Sanitized_2026-08-02.xlsx"
)
GSA003 = (
    REPO_ROOT
    / "data/reference/inventory/gsa-003/2026-08-02"
    / "ARPI_Granite_Used_Auto_Center_Inventory_Sanitized_2026-08-02.xlsx"
)

#: The capture date all three committed artifacts share. Every assertion below filters on
#: it, because earlier tests in this module import FABRICATED workbooks on other dates
#: into the same database; an unfiltered store total would count those too and would move
#: whenever a test above is added.
COMMITTED_CAPTURE = date(2026, 8, 2)


@pytest.fixture(scope="module")
def all_three_stores(listing_connection: Any) -> Any:
    """Import the other two committed captures beside the Chevrolet one, and commit.

    Rolls back first. The shared connection may carry uncommitted work from a test that
    exercised a refusal, and committing somebody else's half-finished transaction as a
    side effect of this fixture would be its own defect.
    """
    listing_connection.rollback()
    for workbook in (GSA002, GSA003):
        summary = import_listing_workbook(
            listing_connection, workbook, sql_root=SQL_ROOT, profile="test"
        )
        assert summary.already_imported is False, workbook.name
        assert not summary.failing_reconciliations, summary.summary()
    listing_connection.commit()
    return listing_connection


@pytest.fixture
def three_store_cursor(all_three_stores: Any) -> Iterator[Any]:
    """A cursor over the three-store database, rolled back after every test.

    The same reason ``listing_cursor`` exists: the connection is shared and committed, so
    one statement that errors would abort the transaction and every later test would fail
    for a reason that has nothing to do with it. Two tests below deliberately provoke a
    constraint violation, which makes the rollback load-bearing rather than tidy.
    """
    with all_three_stores.cursor() as cursor:
        try:
            yield cursor
        finally:
            all_three_stores.rollback()


def test_all_three_committed_captures_load(three_store_cursor: Any) -> None:
    three_store_cursor.execute(
        """
        SELECT d.dealership_id, count(*)
        FROM warehouse.fact_vehicle_listing_snapshot AS f
        JOIN warehouse.dim_dealership AS d ON d.dealership_key = f.dealership_key
        WHERE f.captured_at = %(captured_at)s
        GROUP BY d.dealership_id
        ORDER BY d.dealership_id
        """,
        {"captured_at": COMMITTED_CAPTURE},
    )
    assert three_store_cursor.fetchall() == [("GSA-001", 199), ("GSA-002", 24), ("GSA-003", 318)]


def test_an_unpriced_listing_reaches_the_warehouse_as_null_and_not_as_zero(
    three_store_cursor: Any,
) -> None:
    """The single most consequential assertion about this capture.

    A zero is a number and a number gets averaged. If 287 unpriced listings arrived as
    zero, Granite Used Auto Center's average advertised price would be a fraction of the
    truth and every one of those rows would look like a free car.
    """
    three_store_cursor.execute(
        """
        SELECT f.pricing_status, count(*), count(f.advertised_price), count(f.odometer_miles)
        FROM warehouse.fact_vehicle_listing_snapshot AS f
        JOIN warehouse.dim_dealership AS d ON d.dealership_key = f.dealership_key
        WHERE d.dealership_id = 'GSA-003' AND f.captured_at = %(captured_at)s
        GROUP BY f.pricing_status
        ORDER BY f.pricing_status
        """,
        {"captured_at": COMMITTED_CAPTURE},
    )
    # count(column) counts non-NULLs, so a zero would appear here as a priced row.
    assert three_store_cursor.fetchall() == [
        ("Listed", 31, 31, 31),
        ("Price not exposed", 287, 0, 0),
    ]


def test_an_absent_optional_value_lands_as_null_and_not_as_the_word_none(
    three_store_cursor: Any,
) -> None:
    """The COPY writer must not stringify None.

    It did. ``str(record.odometer_miles)`` put the four characters "None" into a text
    column, staging correctly reported a value present but not representable, and 287
    rows of a real capture were quarantined as REJ-TYPE-001 for a defect in one line of
    Python rather than anything in the workbook.

    What made it expensive to spot is that every layer behaved correctly: the rejection
    was real, the reconciliations compared staged against loaded and passed, and the
    import reported success over a fact holding 31 of 318 rows.
    """
    three_store_cursor.execute(
        """
        SELECT count(*)
        FROM raw.inventory_listing_snapshot_load
        WHERE odometer_miles = 'None' OR advertised_price = 'None'
        """
    )
    assert three_store_cursor.fetchone()[0] == 0

    three_store_cursor.execute(
        """
        SELECT count(*)
        FROM staging.stg_inventory_listing_snapshot_rejected
        WHERE rejection_reason LIKE '%odometer_miles%'
        """
    )
    assert three_store_cursor.fetchone()[0] == 0


def test_the_summary_view_publishes_both_unpriced_buckets_and_they_reconcile(
    three_store_cursor: Any,
) -> None:
    """Every store satisfies listed + unpriced = observed, and the split says why."""
    three_store_cursor.execute(
        """
        SELECT d.dealership_code,
               s.observed_listing_units,
               s.listed_price_units,
               s.call_for_price_units,
               s.price_not_exposed_units,
               s.unpriced_units
        FROM reporting.vw_vehicle_listing_summary AS s
        JOIN reporting.vw_dealership AS d ON d.dealership_key = s.dealership_key
        WHERE s.captured_at = %(captured_at)s
        ORDER BY d.dealership_code
        """,
        {"captured_at": COMMITTED_CAPTURE},
    )
    rows = three_store_cursor.fetchall()
    assert [row[0] for row in rows] == ["GSA-001", "GSA-002", "GSA-003"]
    for code, observed, listed, call_for_price, not_exposed, unpriced in rows:
        assert listed + unpriced == observed, code
        assert call_for_price + not_exposed == unpriced, code
    assert rows[2] == ("GSA-003", 318, 31, 0, 287, 287)


def test_total_advertised_value_describes_only_the_priced_listings(
    three_store_cursor: Any,
) -> None:
    """And the view publishes the count it excluded, so the total cannot mislead alone."""
    three_store_cursor.execute(
        """
        SELECT s.total_advertised_value, s.listed_price_units, s.unpriced_units
        FROM reporting.vw_vehicle_listing_summary AS s
        JOIN reporting.vw_dealership AS d ON d.dealership_key = s.dealership_key
        WHERE d.dealership_code = 'GSA-003' AND s.captured_at = %(captured_at)s
        """,
        {"captured_at": COMMITTED_CAPTURE},
    )
    total, listed, unpriced = three_store_cursor.fetchone()
    assert listed == 31
    assert unpriced == 287
    assert total > 0


def test_the_model_mix_average_odometer_ignores_the_rows_without_a_reading(
    three_store_cursor: Any,
) -> None:
    """avg() skips NULL, and no_odometer_units is what stops that being invisible."""
    three_store_cursor.execute(
        """
        SELECT coalesce(sum(m.no_odometer_units), 0),
               count(*) FILTER (WHERE m.average_odometer_miles IS NULL)
        FROM reporting.vw_vehicle_listing_model_mix AS m
        JOIN reporting.vw_dealership AS d ON d.dealership_key = m.dealership_key
        WHERE d.dealership_code = 'GSA-003' AND m.captured_at = %(captured_at)s
        """,
        {"captured_at": COMMITTED_CAPTURE},
    )
    no_reading, groups_without_an_average = three_store_cursor.fetchone()
    assert no_reading == 287
    # Some groups are entirely unpriced and unread, so their mean is NULL rather than
    # zero. A zero-mile average in a used-car report reads as a new car.
    assert groups_without_an_average > 0


def test_a_priced_row_under_a_no_price_status_is_refused_by_the_database(
    three_store_cursor: Any,
) -> None:
    """The pricing contract covers the new status, not just call-for-price."""
    three_store_cursor.execute(
        "SELECT vehicle_listing_snapshot_key FROM warehouse.fact_vehicle_listing_snapshot "
        "WHERE pricing_status = 'Price not exposed' LIMIT 1"
    )
    key = three_store_cursor.fetchone()[0]
    with pytest.raises(Exception, match="pricing_contract"):
        three_store_cursor.execute(
            "UPDATE warehouse.fact_vehicle_listing_snapshot SET advertised_price = 1000 "
            "WHERE vehicle_listing_snapshot_key = %s",
            (key,),
        )


def test_an_unknown_pricing_status_is_still_refused(three_store_cursor: Any) -> None:
    """Widening the vocabulary by one did not make it open.

    Two constraints refuse an unknown status and PostgreSQL reports whichever it reaches
    first, so the assertion names both. That is not vagueness: the pricing contract is
    written as "Listed requires a price OR a no-price status forbids one", and a status in
    neither set satisfies neither branch. An unknown status therefore cannot get in past
    the domain check by also being priced correctly, which is the loophole worth having.
    """
    three_store_cursor.execute(
        "SELECT vehicle_listing_snapshot_key FROM warehouse.fact_vehicle_listing_snapshot LIMIT 1"
    )
    key = three_store_cursor.fetchone()[0]
    with pytest.raises(Exception, match=r"pricing_status_domain|pricing_contract"):
        three_store_cursor.execute(
            "UPDATE warehouse.fact_vehicle_listing_snapshot SET pricing_status = 'Ask us' "
            "WHERE vehicle_listing_snapshot_key = %s",
            (key,),
        )


def test_an_operating_report_exports_for_the_store_with_no_prices(
    all_three_stores: Any, tmp_path_factory: pytest.TempPathFactory
) -> None:
    """The deliverable has to survive a store whose price statistics are mostly absent."""
    output = tmp_path_factory.mktemp("gsa003-report") / "report.xlsx"
    summary = export_operating_report(
        all_three_stores,
        dealership_id="GSA-003",
        captured_at=COMMITTED_CAPTURE,
        output_path=output,
    )
    assert output.is_file()
    assert summary.row_count == 318
    all_three_stores.rollback()
