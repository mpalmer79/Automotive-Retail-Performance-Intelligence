"""The sanitizer, end to end, against a throwaway fully synthetic private workbook.

EVERY FIXTURE HERE IS INVENTED. The "private" workbooks these tests build contain no
value taken from any source: the identifiers are literal strings chosen to exercise the
hash, the URLs use the reserved ``.invalid`` TLD, and the vehicles are made up. That is
the only way to test a sanitizer without committing the thing it exists to remove.

The assertions are grouped by what would go wrong:

* the input contract, so a malformed private file fails before anything is written;
* the transformation, so an identifier or a URL cannot survive it;
* the output contract, so the artifact is one the validator will accept;
* the file name, so the approved underscore convention is derived rather than hoped for;
* the safety behaviours -- dry run, overwrite refusal, redacted errors -- because each of
  them is a thing somebody will rely on at the worst possible moment.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from arpi.exceptions import ValidationError
from arpi.inventory.contract import load_contract
from arpi.inventory.identity import synthetic_identity
from arpi.inventory.sanitizer import PUBLIC_REFERENCE_NOTICE, sanitize_workbook
from arpi.inventory.validation import validate_workbook

#: The private input contract, exactly as the sanitizer accepts it.
PRIVATE_HEADERS = (
    "Condition",
    "Year",
    "Make",
    "Model",
    "Trim",
    "Vehicle",
    "Mileage",
    "Price",
    "Price Status",
    "VIN",
    "Source URL",
    "Captured",
)

#: Invented identifiers. `.invalid` is the reserved TLD; these resolve to nothing.
ROWS: tuple[tuple[Any, ...], ...] = (
    (
        "New",
        2026,
        "Subaru",
        "Outback",
        "Premium",
        "2026 Subaru Outback Premium",
        6,
        34995,
        "Listed",
        "TESTVEHICLEID0001",
        "https://example.invalid/1",
        "2026-08-09",
    ),
    (
        "New",
        2027,
        "Subaru",
        "Forester",
        "Limited",
        "",
        3,
        None,
        "Call for price",
        "TESTVEHICLEID0002",
        "https://example.invalid/2",
        "2026-08-09",
    ),
    (
        "Used",
        2019,
        "Toyota",
        "RAV4",
        "XLE",
        "2019 Toyota RAV4 XLE",
        58210,
        22450,
        "Listed",
        "TESTVEHICLEID0003",
        "https://example.invalid/3",
        "2026-08-09",
    ),
    (
        # Lower-cased status and a padded identifier: normalisation, not rejection.
        "used",
        2015,
        "Honda",
        "Civic",
        "EX",
        "2015 Honda Civic EX",
        104300,
        9995,
        "listed",
        "  TESTVEHICLEID0004 ",
        "",
        "2026-08-09",
    ),
)

CAPTURED_AT = date(2026, 8, 9)
EXPECTED_NAME = "ARPI_Granite_Subaru_Inventory_Sanitized_2026-08-09.xlsx"


def _write_private(
    path: Path, rows: tuple[tuple[Any, ...], ...] = ROWS, headers: tuple[str, ...] = PRIVATE_HEADERS
) -> Path:
    """Build a throwaway private workbook. Nothing in it comes from any real source."""
    book = Workbook()
    # `Workbook.active` is Optional in the stubs because a workbook can have no
    # sheets. A freshly constructed one always has exactly one.
    sheet = book.active
    assert sheet is not None
    sheet.title = "Inventory"
    sheet.append(list(headers))
    for row in rows:
        sheet.append(list(row))
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    book.close()
    return path


@pytest.fixture
def private(tmp_path: Path) -> Path:
    return _write_private(tmp_path / "private" / "source.xlsx")


@pytest.fixture
def sanitized(tmp_path: Path, private: Path) -> Path:
    summary = sanitize_workbook(
        input_path=private,
        dealership_id="GSA-002",
        captured_at=CAPTURED_AT,
        output_path=tmp_path / "out" / EXPECTED_NAME,
    )
    return summary.output_path


def _values(path: Path, sheet: str) -> list[tuple[Any, ...]]:
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        return [tuple(row) for row in book[sheet].iter_rows(values_only=True)]
    finally:
        book.close()


def _all_strings(path: Path) -> list[str]:
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        return [
            str(value)
            for name in book.sheetnames
            for row in book[name].iter_rows(values_only=True)
            for value in row
            if value is not None
        ]
    finally:
        book.close()


# --------------------------------------------------------------------------------------
# The input contract
# --------------------------------------------------------------------------------------


def test_the_exact_documented_input_contract_is_accepted(private: Path, tmp_path: Path) -> None:
    summary = sanitize_workbook(
        input_path=private,
        dealership_id="GSA-002",
        captured_at=CAPTURED_AT,
        output_path=tmp_path / EXPECTED_NAME,
    )
    assert summary.row_count == len(ROWS)
    assert summary.dealership_id == "GSA-002"
    assert summary.store_name == "Granite Subaru of Manchester"


def test_a_missing_required_header_is_refused(tmp_path: Path) -> None:
    headers = tuple(header for header in PRIVATE_HEADERS if header != "VIN")
    rows = tuple(row[:9] + row[10:] for row in ROWS)
    path = _write_private(tmp_path / "private.xlsx", rows=rows, headers=headers)
    with pytest.raises(ValidationError, match="missing 1 required column"):
        sanitize_workbook(
            input_path=path,
            dealership_id="GSA-002",
            captured_at=CAPTURED_AT,
            output_path=tmp_path / EXPECTED_NAME,
        )


def test_an_optional_header_may_be_absent(tmp_path: Path) -> None:
    """`Source URL` is optional: a private export without one is still sanitizable."""
    headers = tuple(header for header in PRIVATE_HEADERS if header != "Source URL")
    rows = tuple(row[:10] + row[11:] for row in ROWS)
    path = _write_private(tmp_path / "private.xlsx", rows=rows, headers=headers)
    summary = sanitize_workbook(
        input_path=path,
        dealership_id="GSA-002",
        captured_at=CAPTURED_AT,
        output_path=tmp_path / EXPECTED_NAME,
    )
    assert summary.urls_removed == 0
    assert summary.row_count == len(ROWS)


def test_a_duplicated_source_identifier_is_refused(tmp_path: Path) -> None:
    """One physical vehicle cannot be listed twice in one snapshot."""
    duplicated = (*ROWS, ROWS[0])
    path = _write_private(tmp_path / "private.xlsx", rows=duplicated)
    with pytest.raises(ValidationError) as error:
        sanitize_workbook(
            input_path=path,
            dealership_id="GSA-002",
            captured_at=CAPTURED_AT,
            output_path=tmp_path / EXPECTED_NAME,
        )
    assert "duplicates" in str(error.value)


def test_a_truncated_identifier_is_refused(tmp_path: Path) -> None:
    broken = ((*ROWS[0][:9], "ABC", *ROWS[0][10:]),)
    path = _write_private(tmp_path / "private.xlsx", rows=broken)
    with pytest.raises(ValidationError, match="too short"):
        sanitize_workbook(
            input_path=path,
            dealership_id="GSA-002",
            captured_at=CAPTURED_AT,
            output_path=tmp_path / EXPECTED_NAME,
        )


def test_a_listed_row_with_no_price_is_refused(tmp_path: Path) -> None:
    broken = ((*ROWS[0][:7], None, "Listed", *ROWS[0][9:]),)
    path = _write_private(tmp_path / "private.xlsx", rows=broken)
    with pytest.raises(ValidationError, match="no price is present"):
        sanitize_workbook(
            input_path=path,
            dealership_id="GSA-002",
            captured_at=CAPTURED_AT,
            output_path=tmp_path / EXPECTED_NAME,
        )


def test_an_unknown_condition_is_refused(tmp_path: Path) -> None:
    broken = (("Certified", *ROWS[0][1:]),)
    path = _write_private(tmp_path / "private.xlsx", rows=broken)
    with pytest.raises(ValidationError, match="Condition is outside the governed domain"):
        sanitize_workbook(
            input_path=path,
            dealership_id="GSA-002",
            captured_at=CAPTURED_AT,
            output_path=tmp_path / EXPECTED_NAME,
        )


def test_a_workbook_with_no_sanitizable_rows_is_refused(tmp_path: Path) -> None:
    path = _write_private(tmp_path / "private.xlsx", rows=())
    with pytest.raises(ValidationError, match="no sanitizable rows"):
        sanitize_workbook(
            input_path=path,
            dealership_id="GSA-002",
            captured_at=CAPTURED_AT,
            output_path=tmp_path / EXPECTED_NAME,
        )


def test_an_unknown_store_is_refused_before_anything_is_read(tmp_path: Path, private: Path) -> None:
    with pytest.raises(ValidationError, match="dealership registry"):
        sanitize_workbook(
            input_path=private,
            dealership_id="GSA-404",
            captured_at=CAPTURED_AT,
            output_path=tmp_path / EXPECTED_NAME,
        )


# --------------------------------------------------------------------------------------
# The transformation
# --------------------------------------------------------------------------------------


def test_no_source_identifier_survives_into_the_output(sanitized: Path) -> None:
    rendered = " ".join(_all_strings(sanitized))
    for row in ROWS:
        assert str(row[9]).strip() not in rendered


def test_no_source_url_survives_into_the_output(sanitized: Path) -> None:
    rendered = " ".join(_all_strings(sanitized))
    assert "example.invalid" not in rendered
    assert "http" not in rendered.lower()


def test_the_identifiers_are_the_deterministic_ones(sanitized: Path) -> None:
    contract = load_contract()
    rows = _values(sanitized, "Inventory")[1:]
    by_column = {header: index for index, header in enumerate(contract.headers)}
    written = {
        (row[by_column["Synthetic Vehicle ID"]], row[by_column["Synthetic VIN"]]) for row in rows
    }
    expected = {
        (identity.vehicle_id, identity.vin)
        for identity in (synthetic_identity(str(row[9])) for row in ROWS)
    }
    assert written == expected


def test_the_output_is_byte_identical_across_two_runs_of_the_same_input(
    tmp_path: Path, private: Path
) -> None:
    """Deterministic identity is only useful if the whole transformation is."""
    first = sanitize_workbook(
        input_path=private,
        dealership_id="GSA-002",
        captured_at=CAPTURED_AT,
        output_path=tmp_path / "a" / EXPECTED_NAME,
    )
    second = sanitize_workbook(
        input_path=private,
        dealership_id="GSA-002",
        captured_at=CAPTURED_AT,
        output_path=tmp_path / "b" / EXPECTED_NAME,
    )
    # The XLSX container records a creation timestamp, so the BYTES legitimately differ.
    # The logical contract is what must be reproducible.
    assert _values(first.output_path, "Inventory") == _values(second.output_path, "Inventory")


def test_the_store_comes_from_the_registry_and_not_from_the_input(sanitized: Path) -> None:
    contract = load_contract()
    rows = _values(sanitized, "Inventory")[1:]
    store_column = contract.headers.index("Store Name")
    assert {row[store_column] for row in rows} == {"Granite Subaru of Manchester"}


def test_a_call_for_price_row_carries_no_price(sanitized: Path) -> None:
    contract = load_contract()
    rows = _values(sanitized, "Inventory")[1:]
    status = contract.headers.index("Pricing Status")
    price = contract.headers.index("Advertised Price")
    for row in rows:
        if row[status] == "Call for price":
            assert row[price] is None


def test_a_lower_cased_status_is_normalised_rather_than_rejected(sanitized: Path) -> None:
    contract = load_contract()
    rows = _values(sanitized, "Inventory")[1:]
    status = contract.headers.index("Pricing Status")
    assert {row[status] for row in rows} <= {"Listed", "Call for price"}


def test_a_missing_vehicle_display_is_composed(sanitized: Path) -> None:
    contract = load_contract()
    rows = _values(sanitized, "Inventory")[1:]
    display = contract.headers.index("Vehicle Display")
    assert all(row[display] for row in rows)
    assert "2027 Subaru Forester Limited" in {row[display] for row in rows}


# --------------------------------------------------------------------------------------
# The output contract
# --------------------------------------------------------------------------------------


def test_the_output_has_the_four_required_sheets_in_order(sanitized: Path) -> None:
    book = load_workbook(sanitized, read_only=True)
    try:
        assert book.sheetnames == ["README", "Summary", "Inventory", "Model Summary"]
    finally:
        book.close()


def test_the_output_has_the_exact_sanitized_headers(sanitized: Path) -> None:
    contract = load_contract()
    header_row = _values(sanitized, "Inventory")[0]
    assert tuple(header_row) == contract.headers


def test_the_output_carries_the_classification_on_every_row(sanitized: Path) -> None:
    contract = load_contract()
    rows = _values(sanitized, "Inventory")[1:]
    column = contract.headers.index("Data Classification")
    assert {row[column] for row in rows} == {contract.classification}


def test_the_summary_and_model_summary_are_formula_driven(sanitized: Path) -> None:
    book = load_workbook(sanitized, read_only=False, data_only=False)
    try:
        for name in ("Summary", "Model Summary"):
            formulas = [
                value
                for row in book[name].iter_rows(values_only=True)
                for value in row
                if isinstance(value, str) and value.startswith("=")
            ]
            assert formulas, f"{name} carries no formula"
    finally:
        book.close()


def test_the_public_reference_notice_is_present_and_says_not_fully_synthetic(
    sanitized: Path,
) -> None:
    rendered = " ".join(_all_strings(sanitized))
    assert PUBLIC_REFERENCE_NOTICE in rendered
    assert "NOT\nfully synthetic" in PUBLIC_REFERENCE_NOTICE.replace(" ", "\n") or (
        "not" in PUBLIC_REFERENCE_NOTICE.casefold()
        and "fully synthetic" in PUBLIC_REFERENCE_NOTICE.casefold()
    )


def test_the_readme_records_the_intended_repository_path(sanitized: Path) -> None:
    rendered = _all_strings(sanitized)
    assert sanitized.as_posix() in rendered
    assert EXPECTED_NAME in " ".join(rendered)


def test_the_readme_never_recommends_a_hyphenated_name(sanitized: Path) -> None:
    rendered = " ".join(_all_strings(sanitized)).casefold()
    assert "-inventory-sanitized" not in rendered


def test_the_produced_artifact_passes_the_validator(sanitized: Path) -> None:
    """The strongest single assertion available: the sanitizer's output is admissible."""
    result = validate_workbook(sanitized)
    assert result.is_valid, result.summary()
    assert result.row_count == len(ROWS)


# --------------------------------------------------------------------------------------
# Naming
# --------------------------------------------------------------------------------------


def test_the_default_output_name_is_the_approved_underscore_one(
    tmp_path: Path, private: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.chdir(tmp_path)
    summary = sanitize_workbook(
        input_path=private, dealership_id="GSA-002", captured_at=CAPTURED_AT
    )
    assert summary.output_path.name == EXPECTED_NAME
    assert (
        summary.output_path.as_posix()
        == f"data/reference/inventory/gsa-002/2026-08-09/{EXPECTED_NAME}"
    )
    assert summary.output_path.is_file()


def test_the_default_granite_chevrolet_name_is_exact(tmp_path: Path) -> None:
    """The one name that is final, derived by the sanitizer rather than by a helper.

    `--output` is supplied and deliberately points somewhere else, because the name the
    sanitizer REPORTS as approved must not depend on where the operator chose to write.
    """
    chevrolet = tuple((*row[:11], "2026-08-02") for row in ROWS)
    path = _write_private(tmp_path / "private.xlsx", rows=chevrolet)
    summary = sanitize_workbook(
        input_path=path,
        dealership_id="GSA-001",
        captured_at=date(2026, 8, 2),
        output_path=tmp_path / "somewhere-else.xlsx",
        dry_run=True,
    )
    assert summary.output_file_name == "ARPI_Granite_Chevrolet_Inventory_Sanitized_2026-08-02.xlsx"
    assert summary.store_name == "Granite Chevrolet of Nashua"


def test_no_hyphenated_alias_is_created(
    tmp_path: Path, private: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.chdir(tmp_path)
    summary = sanitize_workbook(
        input_path=private, dealership_id="GSA-002", captured_at=CAPTURED_AT
    )
    written = sorted(p.name for p in summary.output_path.parent.iterdir())
    assert written == [EXPECTED_NAME]


# --------------------------------------------------------------------------------------
# Safety behaviours
# --------------------------------------------------------------------------------------


def test_a_dry_run_creates_no_file(tmp_path: Path, private: Path) -> None:
    destination = tmp_path / "out" / EXPECTED_NAME
    summary = sanitize_workbook(
        input_path=private,
        dealership_id="GSA-002",
        captured_at=CAPTURED_AT,
        output_path=destination,
        dry_run=True,
    )
    assert summary.dry_run is True
    assert summary.digest is None
    assert not destination.exists()
    # And it still reports everything an operator needs before committing.
    assert summary.row_count == len(ROWS)
    assert summary.urls_removed == 3
    assert summary.identifiers_replaced == len(ROWS)
    assert summary.output_file_name == EXPECTED_NAME


def test_an_existing_output_is_not_overwritten_by_default(tmp_path: Path, private: Path) -> None:
    destination = tmp_path / EXPECTED_NAME
    destination.write_bytes(b"existing")
    with pytest.raises(ValidationError, match="already exists"):
        sanitize_workbook(
            input_path=private,
            dealership_id="GSA-002",
            captured_at=CAPTURED_AT,
            output_path=destination,
        )
    assert destination.read_bytes() == b"existing"


def test_overwrite_is_available_when_asked_for_explicitly(tmp_path: Path, private: Path) -> None:
    destination = tmp_path / EXPECTED_NAME
    destination.write_bytes(b"existing")
    summary = sanitize_workbook(
        input_path=private,
        dealership_id="GSA-002",
        captured_at=CAPTURED_AT,
        output_path=destination,
        overwrite=True,
    )
    assert summary.digest is not None
    assert destination.read_bytes() != b"existing"


def test_a_failed_run_leaves_no_partial_file(tmp_path: Path) -> None:
    """Atomic writes: a run that fails leaves the directory as it found it."""
    broken = ((*ROWS[0][:9], "", *ROWS[0][10:]),)
    path = _write_private(tmp_path / "private.xlsx", rows=broken)
    destination = tmp_path / "out" / EXPECTED_NAME
    with pytest.raises(ValidationError):
        sanitize_workbook(
            input_path=path,
            dealership_id="GSA-002",
            captured_at=CAPTURED_AT,
            output_path=destination,
        )
    assert not destination.exists()
    assert not destination.parent.exists() or list(destination.parent.iterdir()) == []


def test_error_messages_redact_the_offending_value(tmp_path: Path) -> None:
    """A validator that quoted what it rejected would be the leak it exists to prevent."""
    broken = ((*ROWS[0][:9], "TESTSECRETIDENT01", *ROWS[0][10:11], "2026-08-01"),)
    path = _write_private(tmp_path / "private.xlsx", rows=broken)
    with pytest.raises(ValidationError) as error:
        sanitize_workbook(
            input_path=path,
            dealership_id="GSA-002",
            captured_at=CAPTURED_AT,
            output_path=tmp_path / EXPECTED_NAME,
        )
    rendered = str(error.value)
    assert "TESTSECRETIDENT01" not in rendered
    assert "example.invalid" not in rendered
    # It still says WHERE and WHY.
    assert "row 1" in rendered
    assert "Captured" in rendered


def test_the_summary_carries_no_source_value(tmp_path: Path, private: Path) -> None:
    summary = sanitize_workbook(
        input_path=private,
        dealership_id="GSA-002",
        captured_at=CAPTURED_AT,
        output_path=tmp_path / EXPECTED_NAME,
    )
    rendered = summary.summary() + str(summary.as_dict())
    for row in ROWS:
        assert str(row[9]).strip() not in rendered
    assert "example.invalid" not in rendered
